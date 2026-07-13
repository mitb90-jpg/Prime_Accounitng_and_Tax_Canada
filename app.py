import streamlit as st
import pandas as pd
import io
import re
import json
import os
import datetime
import plotly.graph_objects as go
from supabase import create_client, Client
# ReportLab PDF
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# ---------------- PAGE CONFIG ----------------
st.set_page_config(
    page_title="Smart Transaction Categorizer",
    page_icon="📊",
    layout="wide"
)

# ---------------- DATABASE ----------------
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
supabase: Client = create_client(
    SUPABASE_URL,
    SUPABASE_KEY
)

# ---------------- AUTH ----------------
def signup_form():
    st.subheader("📝 Request Access")
    name = st.text_input("Full Name", key="su_name")
    dob = st.date_input("Date of Birth", key="su_dob",
                         min_value=datetime.date(1950, 1, 1),
                         max_value=datetime.date.today())
    designation = st.text_input("Designation", key="su_designation")
    location = st.text_input("Location of Workplace", key="su_location")
    email = st.text_input("Email", key="su_email")
    password = st.text_input("Password", type="password", key="su_password")

    if st.button("Submit Request"):
        if not all([name, designation, location, email, password]):
            st.warning("Please fill all fields")
            return
        try:
            result = supabase.auth.sign_up(
                {"email": email, "password": password}
            )
            user_id = result.user.id
            supabase.table("profiles").insert({
                "id": user_id,
                "email": email,
                "name": name,
                "dob": str(dob),
                "designation": designation,
                "location": location,
                "role": None,
                "status": "pending"
            }).execute()
            st.success("Request submitted! Wait for admin approval before logging in.")
        except Exception as e:
            st.error(f"Could not submit request: {e}")

def login():

    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"] {
        background: radial-gradient(circle at 15% 15%, #1a3a6b 0%, #0a1128 40%, #05070f 100%) !important;
    }
    [data-testid="stHeader"] {
        background: transparent !important;
    }
    .st-key-login_card_container {
        background: white;
        border-radius: 18px;
        padding: 36px 32px 26px 32px;
        box-shadow: 0 20px 50px rgba(0,0,0,0.35);
        max-width: 440px;
        margin: 40px auto;
    }
    .login-title {
        font-size: 24px;
        font-weight: 700;
        color: #1f4e79;
        text-align: center;
        margin-top: 12px;
        margin-bottom: 2px;
    }
    .login-subtitle {
        font-size: 14px;
        color: #888;
        text-align: center;
        margin-bottom: 20px;
    }
    .st-key-login_card_container .stTextInput input {
        border-radius: 8px !important;
        border: 1.5px solid #e0e0e0 !important;
        padding: 10px 12px !important;
    }
    .st-key-login_card_container div.stButton > button {
        background: linear-gradient(135deg, #1f4e79, #163a5c) !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        padding: 11px !important;
        border: none !important;
        margin-top: 6px !important;
        width: 100% !important;
        color: white !important;
    }
    .st-key-login_card_container div.stButton > button:hover {
        background: linear-gradient(135deg, #163a5c, #0f2a44) !important;
    }
    .st-key-login_card_container [data-baseweb="tab-list"] {
        justify-content: center !important;
        gap: 8px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    with st.container(key="login_card_container"):

        logo_col1, logo_col2, logo_col3 = st.columns([1, 1, 1])
        with logo_col2:
            st.image("Logo.jpeg", width=70)

        st.markdown('<div class="login-title">Prime Accounting and Tax</div>', unsafe_allow_html=True)
        st.markdown('<div class="login-subtitle">Sign in to your account</div>', unsafe_allow_html=True)

        tab_login, tab_signup = st.tabs(["Login", "Request Access"])

        with tab_login:
            email = st.text_input("Email", key="login_email")
            password = st.text_input("Password", type="password", key="login_password")
            if st.button("Login"):
                try:
                    result = supabase.auth.sign_in_with_password(
                        {"email": email, "password": password}
                    )
                    user_id = result.user.id
                    profile = (
                        supabase.table("profiles")
                        .select("*")
                        .eq("id", user_id)
                        .execute()
                    )
                    if not profile.data:
                        st.error("No profile found. Contact admin.")
                        return
                    p = profile.data[0]
                    if p["status"] == "pending":
                        st.warning("Your account is still waiting for admin approval.")
                        return
                    if p["status"] == "rejected":
                        st.error("Your access request was rejected. Contact admin.")
                        return
                    st.session_state.logged_in = True
                    st.session_state.user_email = email
                    st.session_state.user_name = p["name"]
                    st.session_state.role = p["role"]
                    st.rerun()
                except Exception as e:
                    st.error("Invalid email or password")

        with tab_signup:
            signup_form()

def logout():
    supabase.auth.sign_out()
    for key in ["logged_in", "user_email", "user_name", "role"]:
        st.session_state.pop(key, None)
    st.rerun()


def admin_pending_requests():
    st.subheader("🛂 Pending Access Requests")
    
    pending = (
        supabase.table("profiles")
        .select("*")
        .eq("status", "pending")
        .execute()
    ).data
    if not pending:
        st.info("No pending requests")
        return
    for person in pending:
        with st.container(border=True):
            st.write(f"**{person['name']}** ({person['email']})")
            st.write(f"Designation: {person['designation']} | Location: {person['location']} | DOB: {person['dob']}")
            c1, c2, c3 = st.columns(3)
            if c1.button("✅ Approve as Accountant", key=f"acc_{person['id']}"):
                supabase.table("profiles").update(
                    {"role": "accountant", "status": "approved"}
                ).eq("id", person["id"]).execute()
                st.rerun()
            if c2.button("✅ Approve as Newbie", key=f"new_{person['id']}"):
                supabase.table("profiles").update(
                    {"role": "newbie", "status": "approved"}
                ).eq("id", person["id"]).execute()
                st.rerun()
            if c3.button("❌ Reject", key=f"rej_{person['id']}"):
                supabase.table("profiles").update(
                    {"status": "rejected"}
                ).eq("id", person["id"]).execute()
                st.rerun()

# ---------------- VISA CATEGORIZATION DICTIONARIES ----------------

VISA_DEBIT_RULES = {
    "ESSILOR|HCM|NIKON OPTICAL MONTREAL QC|ALCON|CANADIAN OPTICAL|J&J VISION CARE|MCCRAY OPTICAL|OPTICIANS ASSOCIATION|SECURE OPTIONS": "Purchases",

    "REXALL POST OFFICE|ICS|UPS": "Delivery Expenses",

    "Google|CAA MEMBERSHIP|Adobe": "Dues and Subscriptions",

    "Shell|Circle K|COSTCO GAS|Petro": "Fuel",

    "ART OF DENTISTRY|PHARMACY|BAY AND COLLEGE IDA PH|Walgreens|SP REGENEX": "Health and safety",

    "Aviva": "Insurance",

    "Interest|Fee|BMO RETAIL KIOSK": "Interest and Bank charges",

    "KEBABERIE|TST-PLANTA-YORKVILLE|TIM HORTONS|SOUTH ST. BURGER CO.|1 HOTEL|10057 CAVA|ZAROS BAKERY|AC ROUGE|AERA|APROPOS|BASKIN ROBBINS|BMO FOOD & BEVERAGE|BOOSTER JUICE|BROWNS|BUFFALO AND FORT ERIE|CANOE|CASA MEZCAL|CHOTTO MATTE|CHURCH STREET|EAST COAST DONAIR|EL NAHUAL TACO|SQ|STARBUCKS|RITUAL-SOUTH|THE BLAKE HOUSE|HOUSE ON PARLIAMENT|MY DOSA PLACE|SASSAFRAZ|O'GRADY'S RESTAURANT|THE DANISH PASTRY|TRYST WICKEDLILY|FRESH ON CRAWFORD|PAUL'S ROTI SHOP|THE EPICURE SHOP|HAIR OF THE DOG|THE KEG BRAMALEA|SWEET PALACE|GIA|FIRST WATCH|IRENE|WANAS SHAWARMA|TANDOORI FLAME|THE UNDERWING CAMBRIDGE|MALAPARTE|THE CARLU|WHOLE FOODS MARKET": "Meals and Entertainment",

    "WHOLE FOODS MARKET|SHOP LAZZA|FRESHCO|DOLLARAMA|ESSO CIRCLE K|SHOPPERS DRUG MART|LINCOLN ROAD SUPERMARK|INDIAN FROOTLAND|WALMART|KABUL FARMS|ROYAL BLUE GROCERY|RABBA|SARKER GROCERIES|CHERRYCREST ESSO|THE CORNER CONVENIENCE|WELLESLEY CONVENIENCE": "Office Supplies",

    "BARTON PERREIRA|CARL ZEISS VISION|DITA|KERING|LUXOTTICA OF CANADA TORONTO ON|MARCOLIN CANADA|ORGREEN GOLDSMITH|SAFILO|SALT OPTICS|THELIOS|SWEAT AND TONIC": "Purchases",

    "AXIS MEDICAL": "Repairs and Maintenance",

    "UBER|AIRCAN|CITY TOR. FERRY DOCKS|TCKT WEB|CURB|MIAMI AIRP CORONA|COCONUT GROVE|FGTEDCORLANDO|UPE|SIXT|ROAMI|RCL|Prestobay|Kasaliving|Expedia|ENTERPRISE|Canopy|Budget Rent A Car|Avion|Austin Airport|Air BnB|taxi": "Travel Expense",

    "MERCEDES-BENZ|CDN TIRE STORE|SIRIUSXM": "Vehicle Expense",

    "ELEGANCE DRY CLEANERS|HM HENNES|IHUC|LCBO|MEC MOUNTAIN EQUIPMENT|THE BEER STORE|NETFLIX|Air Can|BANANA REPUBLIC|MDHAIR|STORM CROW MANOR|EQUINOX|ATLANTIS EVENTS WEST|FOX & FIDDLE|GABBYS ISABELLA|WOODYS ON CHURCH|WINNERS 379|THE WINE SHOP|THE GREEN CLOSET|TBJ CONCESSIONS|SuitSupply|PALAIS ROYALE|MIXCLOUD LONDON|LA VAPE|HUGO BOSS|HARRY ROSEN|COME RIGHT INN|BUDDIES IN BAD TIMES|BAR+ MUSIC STUDIO|5 DRIVE IN": "Personal Expense",

    "EQUIFAX": "Admin Expenses",

    "Parking|PARK|BIKE SHARE|TORPRKAUT|ERACTOLL|CP 286 POF|ETOLLBGT|CITIPARK": "Parking and Toll",

    "CATALYST": "Promotion",

    "PPL - 21 AVE RD": "Rental",

    "Amazon|COSTCO WHOLESALE|COSTCO|AMZN|JAMESTOWN MILK|STAR FRUIT MARKET|WELLESLEY CONVENIENCE": "Supplies",

    "ROGERS|BELL CANADA|FIDO Home|MOBILEKLINIK|T-MOBILE STORE": "Telephone and Internet",
}

VISA_CREDIT_RULES = {
    "SCOTIABANK PAYMENT|PMT|Telepayments": "Credit Card Payments",
    "Credit Adjustment|Fee": "Other Income",
    "ELEGANCE DRY CLEANERS|REXALL PHARMACY|Costco Wholesale|Air Can|ATLANTIS EVENTS WEST|CANADA WIDE PARKING|AMZN|AC ROUGE|Freshco": "Refund",
}


# ---------------- PROFIT & LOSS STRUCTURE ----------------
# Defines which existing Category names belong in each P&L section, and the
# display order within that section. Categories not listed here fall through
# to "Other/Uncategorized" at the bottom of Expenses. Category names are NOT
# renamed -- they appear in the P&L exactly as produced by the categorization
# rules above.

PL_REVENUE_CATEGORIES = [
    "Revenue",
    "Other Income",
    "Refund",
]

PL_COGS_CATEGORIES = [
    "Purchases",
]

PL_EXPENSE_CATEGORIES = [
    "Meals and Entertainment",
    "Trades and Sub-Contracts",
    "Interest and Bank charges",
    "Supplies",
    "Office Supplies",
    "Dues and Subscriptions",
    "Professional Fee",
    "Admin Expenses",
    "Telephone and Internet",
    "Insurance",
    "Travel Expense",
    "Promotion",
    "Vehicle Expense",
    "Delivery Expenses",
    "Repairs and Maintenance",
    "Parking and Toll",
    "Fuel",
    "Health and safety",
    "Misc Expenses",
    "Rental",
]

# Categories that should never appear in the P&L (internal transfers, not
# real revenue or expense)
PL_EXCLUDED_CATEGORIES = [
    "Credit Card Payments",
]


# ---------------- ROLE HELPERS ----------------
def can_edit():
    return st.session_state.role in ["admin", "accountant"]

def can_delete():
    return st.session_state.role == "admin"

def is_admin():
    return st.session_state.role == "admin"

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    login()
    st.stop()

with st.sidebar:

    st.markdown(
        "<div class='sidebar-brand'>Prime Accounting and Tax Canada</div>",
        unsafe_allow_html=True
    )

    role_badge_class = {
        "admin": "role-badge-admin",
        "accountant": "role-badge-accountant",
        "newbie": "role-badge-newbie"
    }.get(st.session_state.role, "role-badge-newbie")

    user_initial = st.session_state.user_name[0].upper() if st.session_state.user_name else "?"

    st.markdown(
        f"""
        <div class="sidebar-user-card">
            <div class="sidebar-user-avatar">{user_initial}</div>
            <div class="sidebar-user-name">{st.session_state.user_name}</div>
            <span class="role-badge {role_badge_class}">{st.session_state.role}</span>
        </div>
        """,
        unsafe_allow_html=True
    )

    col_refresh, col_logout = st.columns(2)
    with col_refresh:
        if st.button("🔄 Refresh", use_container_width=True):
            st.rerun()
    with col_logout:
        if st.button("🚪 Logout", use_container_width=True):
            logout()

    if st.session_state.role == "admin":
        with st.expander("🛂 Manage Requests"):
            admin_pending_requests()



# ---------------- DATABASE FUNCTIONS ----------------

def clean_import_text(value):
    if pd.isna(value):
        return None
    text = str(value).strip()
    if text == "" or text.lower() == "nan":
        return None
    return text


def clean_import_date(value):
    if pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp, datetime.datetime, datetime.date)):
        try:
            return value.date().isoformat()
        except AttributeError:
            return value.isoformat()
    text = str(value).strip()
    if text == "" or text.lower() == "nan":
        return None
    try:
        from dateutil import parser as date_parser
        parsed = date_parser.parse(text, fuzzy=True)
        return parsed.date().isoformat()
    except Exception:
        return None


def generate_client_code():

    counter = (
        supabase
        .table("client_code_counter")
        .select("last_number")
        .eq("id", 1)
        .execute()
    ).data[0]["last_number"]

    next_number = counter + 1

    supabase.table("client_code_counter") \
        .update({"last_number": next_number}) \
        .eq("id", 1) \
        .execute()

    return f"CL-{next_number:04d}"


def add_client(name, address, contact_number):
    new_code = generate_client_code()
    response = supabase.table("clients").insert(
        {
            "client_name": name,
            "address": address,
            "contact_number": contact_number,
            "created_by": st.session_state.get("user_name", "Unknown"),
            "client_code": new_code,
            "status": "active"
        }
    ).execute()
    return response.data[0]["id"]


def request_deactivate_client(client_id, reason):
    supabase.table("clients") \
        .update({
            "deactivation_requested": True,
            "deactivation_requested_by": st.session_state.get("user_name", "Unknown"),
            "deactivation_requested_reason": reason
        }) \
        .eq("id", client_id) \
        .execute()


def approve_deactivate_client(client_id, reason):
    supabase.table("clients") \
        .update({
            "status": "inactive",
            "deactivation_requested": False,
            "deactivated_by": st.session_state.get("user_name", "Unknown"),
            "deactivation_reason": reason
        }) \
        .eq("id", client_id) \
        .execute()


def reject_deactivate_request(client_id):
    supabase.table("clients") \
        .update({"deactivation_requested": False}) \
        .eq("id", client_id) \
        .execute()


def reactivate_client(client_id):
    supabase.table("clients") \
        .update({
            "status": "active",
            "deactivation_requested": False,
            "deactivated_by": None,
            "deactivation_reason": None
        }) \
        .eq("id", client_id) \
        .execute()



def update_client(
    client_id,
    new_name,
    address,
    contact_number,
    postal_code,
    phone_number,
    email,
    sin_primary,
    sin_spouse,
    dob_primary,
    dob_spouse
):

    supabase.table("clients") \
        .update(
            {
                "client_name": new_name,
                "address": address,
                "contact_number": contact_number,
                "postal_code": postal_code,
                "phone_number": phone_number,
                "email": email,
                "sin_primary": sin_primary,
                "sin_spouse": sin_spouse,
                "dob_primary": str(dob_primary) if dob_primary else None,
                "dob_spouse": str(dob_spouse) if dob_spouse else None
            }
        ) \
        .eq("id", client_id) \
        .execute()



def get_clients():

    response = (
        supabase
        .table("clients")
        .select("id, client_name")
        .order("client_name")
        .execute()
    )

    return response.data


def get_client_details(client_name):

    response = (
        supabase
        .table("clients")
        .select("*")
        .eq("client_name", client_name)
        .execute()
    )

    return response.data[0] if response.data else None

def get_client_details_by_id(client_id):

    response = (
        supabase
        .table("clients")
        .select("*")
        .eq("id", client_id)
        .execute()
    )

    return response.data[0] if response.data else None


# ---------------- INVOICE FUNCTIONS ----------------



def generate_invoice_number():

    today = datetime.date.today()
    prefix = f"INV-PAT-{today.strftime('%y')}"

    response = (
        supabase
        .table("invoices")
        .select("invoice_number")
        .like("invoice_number", f"{prefix}%")
        .order("invoice_number", desc=True)
        .limit(1)
        .execute()
    )

    if response.data:
        last_number = response.data[0]["invoice_number"]
        last_count = int(last_number.split("-")[-1])
        count = last_count + 1
    else:
        count = 1

    return f"{prefix}-{count:04d}"


def add_invoice(
    invoice_number,
    client_name,
    invoice_date,
    due_date,
    description,
    quantity,
    rate,
    amount,
    tax,
    total,
    payment_status,
    received_date
):

    existing = (
        supabase
        .table("invoices")
        .select("invoice_number")
        .eq(
            "invoice_number",
            invoice_number
        )
        .execute()
    )

    if existing.data:
        return

    supabase.table("invoices").insert(
        {
            "invoice_number": invoice_number,
            "client_name": client_name,
            "invoice_date": str(invoice_date),
            "due_date": str(due_date),
            "description": description,
            "quantity": quantity,
            "rate": rate,
            "amount": amount,
            "tax": tax,
            "total": total,
            "payment_status": payment_status,
            "received_date":
                str(received_date)
                if received_date
                else None,
            "created_by": st.session_state.get("user_name", "Unknown")
        }
    ).execute()


def mark_invoice_paid(invoice_number, received_date):

    supabase.table("invoices") \
        .update(
            {
                "payment_status": "Paid",
                "received_date": str(received_date)
            }
        ) \
        .eq("invoice_number", invoice_number) \
        .execute()


def add_invoice_items(invoice_number, items):

    for item in items:

        supabase.table("invoice_items").insert(
            {
                "invoice_number": invoice_number,
                "description": item["Description"],
                "quantity": item["Quantity"],
                "rate": item["Rate"],
                "discount": item.get("Discount", 0),
                "amount": item["Amount"]
            }
        ).execute()


def generate_invoice_pdf(invoice_number):

    invoice = (
        supabase
        .table("invoices")
        .select("*")
        .eq("invoice_number", invoice_number)
        .execute()
    ).data[0]

    client = get_client_details(invoice["client_name"])

    items = (
        supabase
        .table("invoice_items")
        .select("*")
        .eq("invoice_number", invoice_number)
        .execute()
    ).data

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        title="Prime Accounting Invoice",
        topMargin=20,
        bottomMargin=20,
        leftMargin=30,
        rightMargin=30
    )

    styles = getSampleStyleSheet()
    content = []

    TEAL = colors.HexColor("#3a8fa3")
    NAVY = colors.HexColor("#0d2a4a")
    LIGHT_BLUE = colors.HexColor("#bfe1ec")
    LIGHT_BLUE_2 = colors.HexColor("#cfe9f1")

    # ================= TOP BANNER =================

    logo = Image("Logo.jpeg", width=55, height=55)

    banner_data = [[
        Paragraph(
            "<font size=20 color='white'><b>Prime Accounting</b> and Tax</font>",
            styles["Normal"]
        ),
        logo
    ]]

    banner_table = Table(banner_data, colWidths=[420, 90])
    banner_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), NAVY),
        ("BACKGROUND", (1, 0), (1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ("LEFTPADDING", (0, 0), (0, 0), 14),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))

    content.append(banner_table)

    # ================= CONTACT STRIP =================

    contact_data = [[
        Paragraph(
            "<font size=10 color='white'>Toronto, Ontario</font>",
            styles["Normal"]
        ),
        Paragraph(
            "<font size=10 color='white'>Email: info@primetaxes.ca</font>"
            "<br/><font size=10 color='white'>Website: Primetaxes.ca</font>"
            "<br/><font size=10 color='white'>Instagram: Primetaxto</font>",
            styles["Normal"]
        ),
    ]]

    contact_table = Table(contact_data, colWidths=[360, 150])
    contact_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), TEAL),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (0, 0), 14),
        ("RIGHTPADDING", (1, 0), (1, 0), 14),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    content.append(contact_table)
    content.append(Spacer(1, 14))

    # ================= BILL TO / INVOICE INFO =================

    bill_to_text = (
        f"<font size=10><b>Bill To:</b> {invoice['client_name']}</font><br/><br/>"
        f"<font size=10><b>Address:</b> {client.get('address', '') if client else ''}</font>"
    )

    info_text = (
        f"<font size=10><b>Phone:</b> {client.get('contact_number', '') if client else ''}</font><br/><br/>"
        f"<font size=10><b>Invoice #:</b> {invoice['invoice_number']}</font><br/>"
        f"<font size=10><b>Invoice Date:</b> {invoice['invoice_date']}</font><br/>"
        f"<font size=10><b>Due Date:</b> {invoice['due_date']}</font>"
    )

    bill_info_data = [[
        Paragraph(bill_to_text, styles["Normal"]),
        Paragraph(info_text, styles["Normal"]),
    ]]

    bill_info_table = Table(bill_info_data, colWidths=[350, 160])
    bill_info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (0, 0), 0),
    ]))

    content.append(bill_info_table)
    content.append(Spacer(1, 10))

    content.append(
        Paragraph(
            f"<font size=10>&nbsp;&nbsp;<b>Invoice For:</b> {invoice.get('description', '')}</font>",
            styles["Normal"]
        )
    )

    content.append(Spacer(1, 10))

    # ================= ITEMS TABLE =================

    item_data = [["Description", "Qty", "Rate", "Discount", "Price"]]

    for item in items:
        discount_val = item.get("discount", 0) or 0

        item_data.append([
            item["description"],
            str(item["quantity"]),
            f"${item['rate']:,.2f}",
            f"${discount_val:,.2f}" if discount_val else "",
            f"${item['amount']:,.2f}"
        ])

    # pad empty rows so the table always has a consistent height
    min_rows = 8
    while len(item_data) - 1 < min_rows:
        item_data.append(["", "", "", "", "$0.00"])

    item_table = Table(item_data, colWidths=[210, 50, 80, 80, 90])

    item_style = [
        ("BACKGROUND", (0, 0), (-1, 0), LIGHT_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), NAVY),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.white),
    ]

    for row_idx in range(1, len(item_data)):
        bg = LIGHT_BLUE_2 if row_idx % 2 == 1 else LIGHT_BLUE
        item_style.append(("BACKGROUND", (0, row_idx), (-1, row_idx), bg))

    item_table.setStyle(TableStyle(item_style))

    content.append(item_table)
    content.append(Spacer(1, 14))

    # ================= TOTALS BOX =================

    tax_rate_pct = (
        (invoice["tax"] / invoice["amount"] * 100)
        if invoice.get("amount") else 0
    )

    totals_data = [
        ["Invoice Subtotal", f"${invoice['amount']:,.2f}"],
        ["Sales Tax", f"{tax_rate_pct:,.2f}%"],
        ["TOTAL", f"${invoice['total']:,.2f}"],
    ]

    totals_table = Table(totals_data, colWidths=[120, 110])
    totals_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 1), LIGHT_BLUE),
        ("BACKGROUND", (0, 2), (-1, 2), LIGHT_BLUE),
        ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#c0392b")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (1, 0), (1, -1), 10),
    ]))

    # right-align the totals box under the items table
    wrapper = Table(
        [[Paragraph("", styles["Normal"]), totals_table]],
        colWidths=[280, 230]
    )
    wrapper.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))

    content.append(wrapper)
    content.append(Spacer(1, 14))

    # ================= FOOTER =================

    content.append(
        Paragraph(
            "<font size=10><b>Please make payment via e-transfer to: info@primetaxes.ca</b></font>",
            styles["Normal"]
        )
    )

    content.append(Spacer(1, 4))

    content.append(
        Paragraph(
            "<font size=9 color='#3a8fa3'>"
            "Total due in 30 days. Overdue accounts subject to a service charge of 1% per month."
            "</font>",
            styles["Normal"]
        )
    )

    doc.build(content)
    buffer.seek(0)

    return buffer


def get_invoices():

    response = (
        supabase
        .table("invoices")
        .select("*")
        .order(
            "created_at",
            desc=True
        )
        .execute()
    )

    return response.data


def delete_client(client_id):

    supabase.table("clients") \
        .delete() \
        .eq(
            "id",
            client_id
        ) \
        .execute()


def delete_invoice(invoice_number):

    supabase.table("invoices") \
        .delete() \
        .eq(
            "invoice_number",
            invoice_number
        ) \
        .execute()


def void_invoice(invoice_number, reason):

    supabase.table("invoices") \
        .update(
            {
                "is_voided": True,
                "void_reason": reason,
                "voided_by": st.session_state.get("user_name", "Unknown"),
                "voided_at": datetime.datetime.now().isoformat()
            }
        ) \
        .eq("invoice_number", invoice_number) \
        .execute()


def request_void_invoice(invoice_number, reason):

    supabase.table("invoices") \
        .update(
            {
                "void_requested": True,
                "void_requested_by": st.session_state.get("user_name", "Unknown"),
                "void_requested_reason": reason,
                "void_requested_at": datetime.datetime.now().isoformat()
            }
        ) \
        .eq("invoice_number", invoice_number) \
        .execute()


def reject_void_request(invoice_number):

    supabase.table("invoices") \
        .update(
            {
                "void_requested": False
            }
        ) \
        .eq("invoice_number", invoice_number) \
        .execute()



# ---------------- ACCOUNT FUNCTIONS ----------------


def add_account(
    client_id,
    client_name,
    account_name,
    account_type
):

    supabase.table("accounts").insert(
        {
            "client_id": client_id,
            "client_name": client_name,
            "account_name": account_name,
            "account_type": account_type
        }
    ).execute()



def get_accounts(client_id):

    response = (
        supabase
        .table("accounts")
        .select(
            "account_name, account_type"
        )
        .eq(
            "client_id",
            client_id
        )
        .execute()
    )


    return [
        (
            row["account_name"],
            row["account_type"]
        )
        for row in response.data
    ]



# ---------------- FORMAT FUNCTION ----------------
def format_amount(x):
    try:
        if pd.isna(x):
            return ""
        if isinstance(x, (int, float)):
            return f"{x:,.2f}"
        return x
    except:
        return x

# ---------------- VISA STATEMENT PARSER ----------------

def parse_visa_statement(pdf_file):
    """
    Parses an RBC Visa credit card statement PDF.
    Returns a DataFrame with columns: Date, Post Date, Description, Debit, Credit
    """

    import pdfplumber

    def norm(s):
        return re.sub(r"\s+", "", s)

    transactions = []
    current = None

    with pdfplumber.open(pdf_file) as pdf:

        for page in pdf.pages:

            started = False

            words = page.extract_words()

            rows = {}

            for w in words:
                y = round(float(w["top"]), 1)
                rows.setdefault(y, []).append(w)

            for y in sorted(rows):

                line_words = sorted(rows[y], key=lambda x: float(x["x0"]))

                text = " ".join(w["text"] for w in line_words)
                header_norm = norm(text).upper()

                # start of transaction table
                if "TRANSACTION" in header_norm and "POSTING" in header_norm:
                    started = True
                    continue

                if not started:
                    continue

                # stop at total / summary lines (no-space match, since
                # pdfplumber glues these words together with no space)
                if (
                    "TOTALACCOUNTBALANCE" in header_norm
                    or "TIMETOPAY" in header_norm
                    or "INTERESTRATECHART" in header_norm
                ):
                    started = False
                    continue

                # ignore sidebar text (Avion Points, Contact Us, etc.)
                # transaction table content stays left of x=350
                first_x = float(line_words[0]["x0"])

                if first_x >= 350:
                    continue

                first_word = line_words[0]["text"]

                # transaction row starts with month+day glued together, e.g. "JAN11"
                is_date_start = bool(
                    re.match(
                        r"^(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\d{1,2}$",
                        first_word.upper()
                    )
                )

                if is_date_start and len(line_words) >= 2:

                    if current:
                        transactions.append(current)

                    trans_date = line_words[0]["text"]
                    post_date = line_words[1]["text"]

                    current = {
                        "Date": trans_date,
                        "Post Date": post_date,
                        "Description": "",
                        "Amount": ""
                    }

                    for w in line_words[2:]:

                        x = float(w["x0"])
                        value = w["text"]

                        if x >= 300:
                            current["Amount"] += value
                        else:
                            current["Description"] += " " + value

                else:

                    # reference number line (long digits only) -> ignore
                    if re.fullmatch(r"\d{15,}", first_word):
                        continue

                    # otherwise treat as continuation of description
                    if current:
                        current["Description"] += " " + text

            if current:
                transactions.append(current)
                current = None

    df = pd.DataFrame(transactions)

    if df.empty:
        return df

    df["Description"] = df["Description"].str.strip()

    df["Amount"] = (
        df["Amount"]
        .astype(str)
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
    )

    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)

    # Visa convention: negative = payment/credit, positive = purchase (Debit)
    df["Debit"] = df["Amount"].apply(lambda x: x if x > 0 else 0)
    df["Credit"] = df["Amount"].apply(lambda x: abs(x) if x < 0 else 0)

    df = df.drop(columns=["Amount"])

    return df

# ---------------- APPLY VISA CATEGORIES ----------------

def apply_visa_categories(df):
    """
    Applies VISA_DEBIT_RULES and VISA_CREDIT_RULES to a transactions DataFrame.

    Matching uses word boundaries so short keywords (e.g. "FEE") don't
    accidentally match inside longer words (e.g. "COFFEE"). Boundaries are
    only applied on sides of a keyword that start/end with a letter or
    digit, since terms ending in punctuation (e.g. "CO.") don't work
    correctly with a trailing word-boundary marker.

    Empty terms (caused by a trailing "|" in a dictionary entry) are
    automatically skipped.

    If a transaction matches more than one rule, all matching categories
    are joined together and the row is flagged as Needs Review.
    """

    import re

    # Bank descriptions sometimes arrive with no spaces at all
    # (e.g. "ARTOFDENTISTRYTorontoOn"), so we strip ALL whitespace from both
    # the description and each rule keyword before comparing. This makes
    # "ART OF DENTISTRY Toronto On" correctly match "ARTOFDENTISTRYTorontoOn".
    desc_upper = (
        df["Description"].astype(str).str.upper().str.replace(r"\s+", "", regex=True)
    )

    def smart_boundary_term(term):
        term_nospace = re.sub(r"\s+", "", term.upper())
        escaped = re.escape(term_nospace)
        # No leading/trailing \b: once spaces are stripped, merchant names
        # run directly into surrounding text (e.g. "ESSILORCANADAST-LAURENTQC"),
        # so a trailing word-boundary would block legitimate matches.
        return escaped

    def build_pattern(keyword):
        terms = [t.strip() for t in keyword.split("|") if t.strip() != ""]
        if not terms:
            return None
        return "|".join(smart_boundary_term(t) for t in terms)

    categories = [[] for _ in range(len(df))]

    for keyword, category in VISA_DEBIT_RULES.items():

        pattern = build_pattern(keyword)
        if pattern is None:
            continue

        mask = (
            (df["Debit"] > 0) &
            desc_upper.str.contains(pattern, na=False, regex=True)
        )

        for pos in range(len(df)):
            if mask.iloc[pos] and category not in categories[pos]:
                categories[pos].append(category)

    for keyword, category in VISA_CREDIT_RULES.items():

        pattern = build_pattern(keyword)
        if pattern is None:
            continue

        mask = (
            (df["Credit"] > 0) &
            desc_upper.str.contains(pattern, na=False, regex=True)
        )

        for pos in range(len(df)):
            if mask.iloc[pos] and category not in categories[pos]:
                categories[pos].append(category)

    df["Category"] = [" / ".join(c) if c else "" for c in categories]
    df["Needs Review"] = [len(c) > 1 for c in categories]

    return df

# ---------------- TRIANGLE MASTERCARD (7188) STATEMENT PARSER ----------------

def parse_triangle_statement(pdf_file):
    """
    Parses a Triangle World Elite Mastercard (Canadian Tire Bank) statement PDF.
    Returns a DataFrame with columns: Date, Description, Debit, Credit
    Credit = Payments received + Returns and other credits
    Debit  = Purchases + Cash transactions + Fees
    """

    import pdfplumber

    def norm(s):
        return re.sub(r"\s+", "", s)

    MONTHS = "JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC"

    STOP_PHRASES = [
        "CONTINUEDONNEXTPAGE",
        "WAYSTOPAY",
        "NEWADDRESS",
        "DETAILSOFYOURCANADIANTIRESTOREPURCHASES",
        "DETAILSOFYOURINTERESTCHARGES",
        "OTHERDETAILSABOUTYOURACCOUNT",
    ]

    transactions = []
    current = None
    current_section = None  # "credit" or "debit"

    with pdfplumber.open(pdf_file) as pdf:

        for page in pdf.pages:

            in_table = False

            words = page.extract_words()

            rows = {}

            for w in words:
                y = round(float(w["top"]), 1)
                rows.setdefault(y, []).append(w)

            for y in sorted(rows):

                line_words_all = sorted(rows[y], key=lambda x: float(x["x0"]))

                # drop any individual word sitting in the sidebar zone (x >= 380),
                # even if other words on the same line are in the transaction table.
                # This handles cases where a transaction row and an unrelated
                # sidebar paragraph happen to land on the same y-coordinate.
                line_words = [w for w in line_words_all if float(w["x0"]) < 380]

                if not line_words:
                    continue

                text = " ".join(w["text"] for w in line_words)
                header_norm = norm(text).upper()

                # universal stop phrases (footers, disclaimers, unrelated sub-tables)
                if any(p in header_norm for p in STOP_PHRASES):
                    in_table = False
                    current_section = None
                    if current:
                        transactions.append(current)
                        current = None
                    continue

                # sub-label line, doesn't change table state
                if header_norm.startswith("PURCHASES-CARD"):
                    continue

                # section markers
                if header_norm.startswith("PAYMENTSRECEIVED"):
                    current_section = "credit"
                    in_table = False
                    continue

                if header_norm.startswith("RETURNSANDOTHERCREDITS"):
                    current_section = "credit"
                    in_table = False
                    continue

                if header_norm in ("PURCHASES", "PURCHASES(CONTINUED)"):
                    current_section = "debit"
                    in_table = False
                    continue

                if header_norm == "INTERESTCHARGES":
                    current_section = "debit"
                    in_table = False
                    continue

                # Cash transactions section (may wrap onto two lines, e.g.
                # "Cash transactions (includes Balance Transfers" / "if applicable)")
                if header_norm.startswith("CASHTRANSACTIONS"):
                    current_section = "debit"
                    in_table = False
                    continue

                if header_norm == "FEES":
                    current_section = "debit"
                    in_table = False
                    continue

                # table header row -> start capturing (only if we know which section we're in)
                if "TRANSACTION" in header_norm and "POSTING" in header_norm:
                    if current_section is not None:
                        in_table = True
                    continue

                if header_norm.startswith("DATEDATE"):
                    continue

                if not in_table or current_section is None:
                    continue

                # stop at sub-table totals
                if (
                    header_norm.startswith("TOTALPAYMENTSRECEIVED")
                    or header_norm.startswith("TOTALRETURNSANDCREDITS")
                    or header_norm.startswith("TOTALPURCHASES")
                    or header_norm.startswith("TOTALINTERESTCHARGES")
                    or header_norm.startswith("TOTALCASHTRANSACTIONS")
                    or header_norm.startswith("TOTALFEES")
                ):
                    in_table = False
                    current_section = None
                    if current:
                        transactions.append(current)
                        current = None
                    continue

                first_word = line_words[0]["text"]

                is_date_start = bool(
                    re.match(rf"^({MONTHS})$", first_word.upper())
                ) and len(line_words) >= 4

                if is_date_start:

                    if current:
                        transactions.append(current)

                    trans_date = f"{line_words[0]['text']} {line_words[1]['text']}"
                    post_date = f"{line_words[2]['text']} {line_words[3]['text']}"

                    # last word on the line is always the amount;
                    # everything else between dates and amount is the description
                    remaining = line_words[4:]
                    amount_word = remaining[-1]["text"] if remaining else ""
                    desc_words = remaining[:-1]

                    current = {
                        "Date": trans_date,
                        "Post Date": post_date,
                        "Description": " ".join(w["text"] for w in desc_words),
                        "Amount": amount_word,
                        "Section": current_section
                    }

                else:

                    # skip USD foreign-currency conversion lines
                    if re.search(r"USD\s*@", text):
                        continue

                    # continuation of description (wrapped line)
                    if current:
                        current["Description"] += " " + text

            if current:
                transactions.append(current)
                current = None

    df = pd.DataFrame(transactions)

    if df.empty:
        return df

    df["Description"] = df["Description"].str.strip()

    df["Amount"] = (
        df["Amount"]
        .astype(str)
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
    )

    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)

    # Section already tells us credit vs debit directly (no sign-guessing needed)
    df["Debit"] = df.apply(lambda r: abs(r["Amount"]) if r["Section"] == "debit" else 0, axis=1)
    df["Credit"] = df.apply(lambda r: abs(r["Amount"]) if r["Section"] == "credit" else 0, axis=1)

    df = df.drop(columns=["Amount", "Section"])

    return df


# ---------------- CSS ----------------
st.markdown("""
<style>
div.stDownloadButton > button {
    width: 100%;
    background-color: #1f4e79;
    color: white;
    font-size: 16px;
    font-weight: bold;
    padding: 12px;
    border-radius: 10px;
}
div.stDownloadButton > button:hover {
    background-color: #163a5c;
}
div.stButton > button {
    width: 100%;
    background-color: #1f4e79;
    color: white;
    font-size: 16px;
    font-weight: bold;
    padding: 12px;
    border-radius: 10px;
}
div.stButton > button:hover {
    background-color: #163a5c;
}
section[data-testid="stSidebar"] {
    background-color: #f0f2f5;
    border-right: 1px solid #e5e7eb;
}
section[data-testid="stSidebar"] * {
    color: #333333;
}

section[data-testid="stSidebar"] .sidebar-user-card,
section[data-testid="stSidebar"] .sidebar-user-card * {
    color: white !important;
}
section[data-testid="stSidebar"] .stRadio label {
    font-size: 16px;
    padding: 10px 14px;
    border-radius: 8px;
    margin-bottom: 4px;
    display: block;
}
section[data-testid="stSidebar"] .stRadio label[data-checked="true"] {
    background-color: #eaf1fb;
}
section[data-testid="stSidebar"] .stRadio label[data-checked="true"] p {
    color: #1f4e79 !important;
    font-weight: bold;
}
section[data-testid="stSidebar"] div.stButton > button p {
    color: white !important;
    font-weight: bold;
}

/* ---------------- FANCY FILE UPLOADERS (different colors per card) ---------------- */

div[data-testid="stFileUploader"] section {
    border: none;
    background-color: transparent;
}

div[data-testid="stFileUploader"] button {
    border-radius: 10px !important;
    font-weight: bold !important;
    font-size: 16px !important;
    padding: 10px 20px !important;
    border: none !important;
    color: white !important;
}

/* Scotia - blue */
.st-key-uploader_scotia_container div[data-testid="stFileUploader"] {
    border-radius: 14px;
    padding: 14px;
    background-color: #e8f1fb;
    box-shadow: 0 4px 15px rgba(0,0,0,0.08);
}
.st-key-uploader_scotia_container div[data-testid="stFileUploader"] button {
    background-color: #1f4e79 !important;
}
.st-key-uploader_scotia_container div[data-testid="stFileUploader"] button:hover {
    background-color: #163a5c !important;
}

/* Visa - amber */
.st-key-uploader_visa_container div[data-testid="stFileUploader"] {
    border-radius: 14px;
    padding: 14px;
    background-color: #fef3e0;
    box-shadow: 0 4px 15px rgba(0,0,0,0.08);
}
.st-key-uploader_visa_container div[data-testid="stFileUploader"] button {
    background-color: #b8762f !important;
}
.st-key-uploader_visa_container div[data-testid="stFileUploader"] button:hover {
    background-color: #96601f !important;
}

/* Triangle Mastercard - teal */
.st-key-uploader_triangle_container div[data-testid="stFileUploader"] {
    border-radius: 14px;
    padding: 14px;
    background-color: #e3f4f4;
    box-shadow: 0 4px 15px rgba(0,0,0,0.08);
}
.st-key-uploader_triangle_container div[data-testid="stFileUploader"] button {
    background-color: #2f7d7d !important;
}
.st-key-uploader_triangle_container div[data-testid="stFileUploader"] button:hover {
    background-color: #235e5e !important;
}

/* ---------------- SIDEBAR: DARK GLOW THEME ---------------- */

section[data-testid="stSidebar"] {
    background: radial-gradient(circle at 0% 0%, #1a3a6b 0%, #0a1128 35%, #05070f 100%) !important;
    border-right: none !important;
}

section[data-testid="stSidebar"] * {
    color: #cbd5e1 !important;
}

.sidebar-brand {
    padding: 6px 0 16px 0;
    font-size: 15px;
    font-weight: 700;
    color: #ffffff !important;
    letter-spacing: 0.3px;
    line-height: 1.3;
}

.sidebar-user-card {
    padding: 14px 0 18px 0;
    margin-bottom: 8px;
    border-bottom: 1px solid rgba(255,255,255,0.08);
}

.sidebar-user-avatar {
    width: 38px;
    height: 38px;
    border-radius: 50%;
    background: linear-gradient(135deg, #3b82f6, #1e3a8a);
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 15px;
    font-weight: 700;
    color: white !important;
    float: left;
    margin-right: 12px;
}

.sidebar-user-name {
    font-size: 14.5px;
    font-weight: 600;
    color: #ffffff !important;
    margin-top: 4px;
    margin-bottom: 2px;
}

.role-badge {
    font-size: 10.5px;
    font-weight: 600;
    letter-spacing: 0.8px;
    text-transform: uppercase;
}

.role-badge-admin { color: #f0c674 !important; }
.role-badge-accountant { color: #5dc4d6 !important; }
.role-badge-newbie { color: #8b96a3 !important; }

.sidebar-menu-label {
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 1.5px;
    color: #64748b !important;
    margin: 4px 0 8px 4px;
    text-transform: uppercase;
}

/* ---------------- SIDEBAR NAV: dark theme active state ---------------- */

section[data-testid="stSidebar"] div.stButton > button {
    background-color: transparent !important;
    color: #cbd5e1 !important;
    font-weight: 500 !important;
    text-align: left !important;
    box-shadow: none !important;
    border-radius: 8px !important;
    padding: 10px 14px !important;
    justify-content: flex-start !important;
    display: flex !important;
}

section[data-testid="stSidebar"] div.stButton > button p {
    text-align: left !important;
    width: 100% !important;
}

section[data-testid="stSidebar"] div.stButton > button:hover {
    background-color: rgba(255,255,255,0.06) !important;
}

section[data-testid="stSidebar"] div.stButton > button[kind="primary"] {
    background-color: rgba(59,130,246,0.18) !important;
    color: #ffffff !important;
    font-weight: 600 !important;
    border-left: 3px solid #3b82f6 !important;
    border-radius: 6px !important;
}

/* ---------------- SIDEBAR STATS: dark theme ---------------- */

.sidebar-stats-card {
    padding-top: 14px;
    margin-top: 10px;
    border-top: 1px solid rgba(255,255,255,0.08);
}

.sidebar-stat-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 7px 0;
    font-size: 13.5px;
}

.sidebar-stat-label {
    color: #8b96a3 !important;
}

.sidebar-stat-value {
    font-weight: 700;
    color: #5dc4d6 !important;
    font-size: 15px;
}

/* ---------------- MAIN APP BACKGROUND ---------------- */

[data-testid="stAppViewContainer"] {
    background: linear-gradient(160deg, #f4f7fb 0%, #eef2f9 45%, #e8eef7 100%) !important;
}

[data-testid="stHeader"] {
    background: transparent !important;
}

.main .block-container {
    padding-top: 2rem;
}

/* ---------------- PROFESSIONAL DATA TABLES ---------------- */

[data-testid="stDataFrame"] {
    border-radius: 12px !important;
    overflow: hidden !important;
    box-shadow: 0 2px 10px rgba(0,0,0,0.06) !important;
    border: 1px solid #e5e7eb !important;
}

[data-testid="stDataFrame"] div[data-testid="stDataFrameResizable"] {
    border-radius: 12px !important;
}

[data-testid="stElementToolbar"] {
    background: transparent !important;
}

/* ---------------- LOGIN PAGE ---------------- */

.login-page-wrapper [data-testid="stAppViewContainer"] {
    background: radial-gradient(circle at 15% 15%, #1a3a6b 0%, #0a1128 40%, #05070f 100%) !important;
}

.login-card {
    background: white;
    border-radius: 18px;
    padding: 40px 36px 30px 36px;
    box-shadow: 0 20px 50px rgba(0,0,0,0.35);
    max-width: 440px;
    margin: 0 auto;
}

.login-title {
    font-size: 24px;
    font-weight: 700;
    color: #1f4e79;
    text-align: center;
    margin-top: 14px;
    margin-bottom: 2px;
}

.login-subtitle {
    font-size: 14px;
    color: #888;
    text-align: center;
    margin-bottom: 22px;
}

.login-page-wrapper [data-testid="stForm"],
.login-page-wrapper div.stTabs {
    max-width: 440px;
    margin: 0 auto;
}

.login-page-wrapper .stTextInput input {
    border-radius: 8px !important;
    border: 1.5px solid #e0e0e0 !important;
    padding: 10px 12px !important;
}

.login-page-wrapper .stTextInput input:focus {
    border-color: #3b82f6 !important;
    box-shadow: 0 0 0 2px rgba(59,130,246,0.15) !important;
}

.login-page-wrapper div.stButton > button {
    background: linear-gradient(135deg, #1f4e79, #163a5c) !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    padding: 11px !important;
    border: none !important;
    margin-top: 8px !important;
}

.login-page-wrapper div.stButton > button:hover {
    background: linear-gradient(135deg, #163a5c, #0f2a44) !important;
}

.login-page-wrapper [data-baseweb="tab-list"] {
    justify-content: center !important;
    gap: 8px !important;
}
</style>
""", unsafe_allow_html=True)


# ---------------- APP MENU ----------------
st.sidebar.markdown("---")
if "page" not in st.session_state:
    st.session_state.page = "🏠 Dashboard"

nav_items = [
    "🏠 Dashboard",
    "👥 Clients",
    "🧾 Sales",
    "📄 Invoice History",    
    "📊 Reports"
]

st.sidebar.markdown("<div class='sidebar-menu-label'>Menu</div>", unsafe_allow_html=True)

for item in nav_items:
    is_active = (item == st.session_state.page)
    if st.sidebar.button(
        item,
        key=f"nav_{item}",
        use_container_width=True,
        type="primary" if is_active else "secondary"
    ):
        st.session_state.page = item
        st.rerun()

page = st.session_state.page

st.sidebar.markdown("---")

total_clients = len(get_clients())

all_invoices = get_invoices()

unpaid_count = len([
    inv for inv in all_invoices
    if inv["payment_status"] == "Unpaid"
])

st.sidebar.markdown(
    f"""
    <div class="sidebar-stats-card">
        <div class="sidebar-stat-row">
            <span class="sidebar-stat-label">👥 Total Clients</span>
            <span class="sidebar-stat-value">{total_clients}</span>
        </div>
        <div class="sidebar-stat-row">
            <span class="sidebar-stat-label">📌 Unpaid Invoices</span>
            <span class="sidebar-stat-value">{unpaid_count}</span>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)

# ================= CLIENT PAGE =================

if page == "👥 Clients":

    st.title("👥 Client Management")

    if "client_category" not in st.session_state:
        st.session_state.client_category = "Personal Taxes"

    cat_col1, cat_col2 = st.columns(2)

    with cat_col1:
        if st.button(
            "👤 Personal Taxes",
            use_container_width=True,
            type="primary" if st.session_state.client_category == "Personal Taxes" else "secondary"
        ):
            st.session_state.client_category = "Personal Taxes"
            st.rerun()

    with cat_col2:
        if st.button(
            "🏢 Corporate Taxes",
            use_container_width=True,
            type="primary" if st.session_state.client_category == "Corporate Taxes" else "secondary"
        ):
            st.session_state.client_category = "Corporate Taxes"
            st.rerun()

    st.divider()

    if st.session_state.client_category == "Corporate Taxes":

        st.info("🏗️ Corporate Tax client management is coming soon. We'll build this out next.")

        st.stop()

    if is_admin():
        with st.expander("📥 Bulk Import Clients (Excel)"):

            st.write(
                "Excel columns expected: Sr No., Client Name, SIN Primary, SIN Spouse, "
                "Address, Postal code, dob primary, dob spouse, Phone Number, E-Mail"
            )

            import_file = st.file_uploader(
                "Choose Excel file",
                type=["xlsx"],
                key="client_import_uploader"
            )

            if import_file is not None:

                import_df = pd.read_excel(import_file, header=2)
                import_df = import_df.dropna(subset=["Client Name"])

                st.write(f"Found **{len(import_df)}** client rows ready to import")

                st.dataframe(import_df, use_container_width=True, hide_index=True)

                if st.button("✅ Confirm Import", key="confirm_client_import"):

                    imported_count = 0

                    for _, row in import_df.iterrows():

                        name = clean_import_text(row.get("Client Name"))

                        if not name:
                            continue

                        new_code = generate_client_code()
                        phone = clean_import_text(row.get("Phone Number"))

                        supabase.table("clients").insert({
                            "client_name": name,
                            "address": clean_import_text(row.get("Address")),
                            "contact_number": phone,
                            "created_by": st.session_state.get("user_name", "Unknown"),
                            "client_code": new_code,
                            "status": "active",
                            "sin_primary": clean_import_text(row.get("SIN Primary")),
                            "sin_spouse": clean_import_text(row.get("SIN Spouse")),
                            "postal_code": clean_import_text(row.get("Postal code")),
                            "dob_primary": clean_import_date(row.get("dob primary")),
                            "dob_spouse": clean_import_date(row.get("dob spouse")),
                            "phone_number": phone,
                            "email": clean_import_text(row.get("E-Mail"))
                        }).execute()

                        imported_count += 1

                    st.success(f"✅ Imported {imported_count} clients successfully")
                    st.rerun()

    if "clients_active_tab" not in st.session_state:
        st.session_state.clients_active_tab = "All Clients"

    tab_col1, tab_col2, tab_col3 = st.columns(3)

    with tab_col1:
        if st.button("📋 All Clients", use_container_width=True):
            st.session_state.clients_active_tab = "All Clients"

    with tab_col2:
        if can_edit():
            if st.button("➕ Add Client", use_container_width=True):
                st.session_state.clients_active_tab = "Add Client"

    with tab_col3:
        if st.button("👤 Client Profile", use_container_width=True):
            st.session_state.clients_active_tab = "Client Profile"

    st.divider()

    clients = get_clients()

    client_labels = [
        f"{c['client_name']} (ID: {c['id']})"
        for c in clients
    ]

    def get_client_id_from_label(label):
        return int(label.split("ID: ")[1].rstrip(")"))

    if st.session_state.clients_active_tab == "Add Client":

        st.subheader("Add New Client")

        client_name = st.text_input(
            "Client Name"
        )

        client_address = st.text_input(
            "Address"
        )

        client_contact = st.text_input(
            "Business Number"
        )

        st.markdown("**Account Details (optional)**")

        if "new_client_accounts" not in st.session_state:
            st.session_state.new_client_accounts = []

        new_account_name = st.text_input(
            "Account Name",
            placeholder="Example: Scotia Bank",
            key="new_client_account_name"
        )

        new_account_type = st.selectbox(
            "Account Type",
            ["Bank Account", "Credit Card"],
            key="new_client_account_type"
        )

        if st.button("➕ Add Account", key="add_pending_account"):

            if new_account_name.strip():

                st.session_state.new_client_accounts.append(
                    {
                        "Account Name": new_account_name,
                        "Account Type": new_account_type
                    }
                )

                st.rerun()

        if st.session_state.new_client_accounts:

            st.dataframe(
                pd.DataFrame(st.session_state.new_client_accounts),
                use_container_width=True,
                hide_index=True
            )


        if st.button("➕ Add Client"):

            if client_name.strip():

                new_client_id = add_client(client_name, client_address, client_contact)

                for acc in st.session_state.new_client_accounts:

                    add_account(
                        new_client_id,
                        client_name,
                        acc["Account Name"],
                        acc["Account Type"]
                    )

                st.session_state.new_client_accounts = []

                st.success(
                    "Client Added Successfully"
                )

                st.rerun()



    if st.session_state.clients_active_tab == "All Clients":

        header_col, search_col = st.columns([2, 2])

        with header_col:
            st.subheader("Existing Clients")

        if clients:

            all_clients_data = (
                supabase
                .table("clients")
                .select("*")
                .order("client_name")
                .execute()
            ).data

            client_df = pd.DataFrame(all_clients_data)

            client_df = client_df.rename(columns={
                "id": "Client ID",
                "client_code": "Client Code",
                "client_name": "Client Name",
                "address": "Address",
                "contact_number": "Contact Number",
                "created_by": "Created By",
                "status": "Status"
            })

            if "Status" in client_df.columns:
                client_df["Status"] = client_df["Status"].apply(
                    lambda x: "🟢 Active" if x == "active" else "🔴 Inactive"
                )

            with search_col:
                client_search = st.text_input(
                    "🔍 Search clients",
                    key="client_search_box",
                    placeholder="Search by name, address, or contact..."
                )

            if client_search.strip():
                mask = client_df.apply(
                    lambda row: client_search.lower() in str(row).lower(),
                    axis=1
                )
                filtered_df = client_df[mask].reset_index(drop=True)
            else:
                filtered_df = client_df

            filtered_df = filtered_df.drop(columns=["Sr. No"], errors="ignore")
            filtered_df.insert(0, "Sr. No", range(1, len(filtered_df) + 1))

            st.markdown(
                f"<p style='color:#888; font-size:13.5px; margin-bottom:8px;'>"
                f"Showing <b>{len(filtered_df)}</b> of <b>{len(client_df)}</b> clients</p>",
                unsafe_allow_html=True
            )

            st.dataframe(
                filtered_df,
                hide_index=True,
                use_container_width=True,
                height=min(70 + len(filtered_df) * 36, 600),
                column_config={
                    "Sr. No": st.column_config.NumberColumn("Sr. No", width="small"),
                    "Client ID": st.column_config.NumberColumn("Client ID", width="small"),
                    "Client Code": st.column_config.TextColumn("Client Code", width="small"),
                    "Client Name": st.column_config.TextColumn("Client Name", width="medium"),
                    "Address": st.column_config.TextColumn("Address", width="large"),
                    "Contact Number": st.column_config.TextColumn("Contact Number", width="medium"),
                    "Created By": st.column_config.TextColumn("Created By", width="medium"),
                    "Status": st.column_config.TextColumn("Status", width="small"),
                }
            )

            client_excel = io.BytesIO()

            with pd.ExcelWriter(client_excel, engine="openpyxl") as writer:
                client_df.to_excel(writer, index=False, sheet_name="Clients")

            client_excel.seek(0)

            st.download_button(
                "⬇️ Export Clients Database",
                data=client_excel,
                file_name="Clients_Database.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        else:

            st.info(
                "No clients added yet"
            )

        st.divider()

        if can_delete():
            st.subheader("Delete Client")

            if clients:

                delete_client_label = st.selectbox(
                    "Select Client",
                    options=["Select Client"] + client_labels,
                    index=0,
                    key="delete_client_dropdown"
                )

                if "confirm_delete" not in st.session_state:
                    st.session_state.confirm_delete = False

                if st.button("🗑️ Delete Client"):

                    if delete_client_label != "Select Client":
                        st.session_state.confirm_delete = True
                    else:
                        st.warning("Please select a client first")

                if st.session_state.confirm_delete:

                    st.warning(
                        f"⚠️ Are you sure you want to delete '{delete_client_label}'?"
                    )

                    c1, c2 = st.columns(2)

                    with c1:
                        if st.button("✅ Yes, Delete"):
                            delete_id = get_client_id_from_label(delete_client_label)
                            delete_client(delete_id)
                            st.session_state.confirm_delete = False
                            st.success("Client Deleted")
                            st.rerun()

                    with c2:
                        if st.button("❌ Cancel"):
                            st.session_state.confirm_delete = False
                            st.rerun()

            else:

                st.info(
                    "No clients available to delete"
                )

    if st.session_state.clients_active_tab == "Client Profile":

        st.subheader("👤 Client Profile")

        if clients:

            profile_client_label = st.selectbox(
                "Select Client to View Profile",
                ["Select Client"] + client_labels,
                key="profile_client_select"
            )

            if profile_client_label != "Select Client":

                profile_client_id = get_client_id_from_label(profile_client_label)

                details = get_client_details_by_id(profile_client_id)

                status_display = "🟢 Active" if details.get("status", "active") == "active" else "🔴 Inactive"

                st.write(f"**Client Code:** {details.get('client_code', 'N/A')}")
                st.write(f"**Name:** {details['client_name']}")
                st.write(f"**Status:** {status_display}")
                st.write(f"**Address:** {details.get('address', '') or '—'}")
                st.write(f"**Postal Code:** {details.get('postal_code', '') or '—'}")
                st.write(f"**Phone Number:** {details.get('phone_number') or details.get('contact_number') or '—'}")
                st.write(f"**Email:** {details.get('email', '') or '—'}")
                st.write(f"**SIN (Primary):** {details.get('sin_primary', '') or '—'}")
                st.write(f"**SIN (Spouse):** {details.get('sin_spouse', '') or '—'}")
                st.write(f"**Date of Birth (Primary):** {details.get('dob_primary', '') or '—'}")
                st.write(f"**Date of Birth (Spouse):** {details.get('dob_spouse', '') or '—'}")

                if can_edit():

                    st.divider()

                    st.markdown("**🔄 Client Status**")

                    is_inactive = details.get("status") == "inactive"
                    is_pending_deactivation = details.get("deactivation_requested")

                    if is_inactive:

                        st.warning(
                            f"This client is INACTIVE.\n\n"
                            f"Reason: {details.get('deactivation_reason', '')}\n\n"
                            f"Deactivated by: {details.get('deactivated_by', '')}"
                        )

                        if is_admin():
                            if st.button("🟢 Reactivate Client", key="reactivate_client_btn"):
                                reactivate_client(details["id"])
                                st.success("Client reactivated")
                                st.rerun()

                    elif is_pending_deactivation and is_admin():

                        st.info(
                            f"⏳ Deactivation request pending your approval\n\n"
                            f"Requested by: {details.get('deactivation_requested_by', '')}\n\n"
                            f"Reason: {details.get('deactivation_requested_reason', '')}"
                        )

                        c1, c2 = st.columns(2)

                        with c1:
                            if st.button("✅ Approve & Deactivate", key="approve_deactivate_btn"):
                                approve_deactivate_client(
                                    details["id"],
                                    details.get("deactivation_requested_reason", "")
                                )
                                st.success("Client marked Inactive")
                                st.rerun()

                        with c2:
                            if st.button("❌ Reject Request", key="reject_deactivate_btn"):
                                reject_deactivate_request(details["id"])
                                st.info("Request rejected")
                                st.rerun()

                    elif is_pending_deactivation and not is_admin():

                        st.info(
                            f"⏳ Deactivation request submitted, waiting for admin approval.\n\n"
                            f"Reason: {details.get('deactivation_requested_reason', '')}"
                        )

                    elif is_admin():

                        deactivate_reason = st.text_area(
                            "Reason for deactivating this client (required)",
                            key="deactivate_reason_input"
                        )

                        if st.button("🔴 Deactivate Client", key="deactivate_client_btn"):
                            if deactivate_reason.strip():
                                approve_deactivate_client(details["id"], deactivate_reason)
                                st.success("Client marked Inactive")
                                st.rerun()
                            else:
                                st.warning("Please enter a reason")

                    else:

                        deactivate_request_reason = st.text_area(
                            "Reason for requesting deactivation (required)",
                            key="deactivate_request_reason_input"
                        )

                        if st.button("📨 Request Deactivation", key="request_deactivate_btn"):
                            if deactivate_request_reason.strip():
                                request_deactivate_client(details["id"], deactivate_request_reason)
                                st.success("Deactivation request sent to admin")
                                st.rerun()
                            else:
                                st.warning("Please enter a reason")

                if is_admin():
                    with st.expander("✏️ Edit Client Details"):

                        edit_name = st.text_input(
                            "Client Name",
                            value=details['client_name'],
                            key=f"edit_client_name_{details['id']}"
                        )

                        edit_address = st.text_input(
                            "Address",
                            value=details.get('address') or '',
                            key=f"edit_client_address_{details['id']}"
                        )

                        edit_contact = st.text_input(
                            "Contact Number",
                            value=details.get('contact_number') or '',
                            key=f"edit_client_contact_{details['id']}"
                        )

                        edit_postal = st.text_input(
                            "Postal Code",
                            value=details.get('postal_code') or '',
                            key=f"edit_client_postal_{details['id']}"
                        )

                        edit_phone = st.text_input(
                            "Phone Number",
                            value=details.get('phone_number') or '',
                            key=f"edit_client_phone_{details['id']}"
                        )

                        edit_email = st.text_input(
                            "Email",
                            value=details.get('email') or '',
                            key=f"edit_client_email_{details['id']}"
                        )

                        edit_sin_primary = st.text_input(
                            "SIN (Primary)",
                            value=details.get('sin_primary') or '',
                            key=f"edit_client_sin_primary_{details['id']}"
                        )

                        edit_sin_spouse = st.text_input(
                            "SIN (Spouse)",
                            value=details.get('sin_spouse') or '',
                            key=f"edit_client_sin_spouse_{details['id']}"
                        )

                        existing_dob_primary = None
                        if details.get('dob_primary'):
                            try:
                                existing_dob_primary = datetime.date.fromisoformat(str(details['dob_primary'])[:10])
                            except ValueError:
                                existing_dob_primary = None

                        edit_dob_primary = st.date_input(
                            "Date of Birth (Primary)",
                            value=existing_dob_primary,
                            min_value=datetime.date(1900, 1, 1),
                            max_value=datetime.date.today(),
                            key=f"edit_client_dob_primary_{details['id']}"
                        )

                        existing_dob_spouse = None
                        if details.get('dob_spouse'):
                            try:
                                existing_dob_spouse = datetime.date.fromisoformat(str(details['dob_spouse'])[:10])
                            except ValueError:
                                existing_dob_spouse = None

                        edit_dob_spouse = st.date_input(
                            "Date of Birth (Spouse)",
                            value=existing_dob_spouse,
                            min_value=datetime.date(1900, 1, 1),
                            max_value=datetime.date.today(),
                            key=f"edit_client_dob_spouse_{details['id']}"
                        )

                        if st.button("💾 Save Changes", key="save_client_edit"):

                            update_client(
                                profile_client_id,
                                edit_name,
                                edit_address,
                                edit_contact,
                                edit_postal,
                                edit_phone,
                                edit_email,
                                edit_sin_primary,
                                edit_sin_spouse,
                                edit_dob_primary,
                                edit_dob_spouse
                            )

                            st.success("Client Updated Successfully")

                            st.rerun()

                st.divider()

                st.subheader("🏦 Accounts")

                if can_edit():
                    account_name = st.text_input(
                        "Account Name",
                        placeholder="Example: Scotia Bank",
                        key="profile_account_name"
                    )

                    account_type = st.selectbox(
                        "Account Type",
                        ["Bank Account", "Credit Card"],
                        key="profile_account_type"
                    )

                    if st.button("➕ Add Account", key="profile_add_account"):

                        if account_name.strip():
                            add_account(details["id"], details['client_name'], account_name, account_type)
                            st.success("Account Added Successfully")
                            st.rerun()

                accounts = get_accounts(details["id"])

                if accounts:

                    account_df = pd.DataFrame(
                        accounts,
                        columns=["Account Name", "Account Type"]
                    )

                    st.dataframe(
                        account_df,
                        use_container_width=True,
                        hide_index=True
                    )

                else:

                    st.info("No accounts added for this client")


# ================= SALES PAGE =================

if page == "🧾 Sales":

    if "sales_form_version" not in st.session_state:
        st.session_state.sales_form_version = 0

    v = st.session_state.sales_form_version

    col_title, col_refresh = st.columns([4, 1])

    with col_title:
        st.title("🧾 Sales & Invoice Management")

    with col_refresh:
        if st.button("🔄 Refresh Page", key="manual_refresh_button"):

            st.session_state.invoice_items = []
            st.session_state.current_invoice_number = generate_invoice_number()
            st.session_state.sales_form_version += 1

            st.rerun()

    if can_edit():
        st.subheader("Create Invoice")



        # ---------- INVOICE BASIC DETAILS ----------

        col1, col2 = st.columns(2)

        with col1:

            sales_clients = get_clients()

            sales_client_labels = [
                f"{c['client_name']} (ID: {c['id']})"
                for c in sales_clients
            ]

            customer_label = st.selectbox(
                "Customer",
                ["Select Client"] + sales_client_labels,
                key=f"sales_customer_label_{v}"
            )

            customer_name = (
                customer_label.split(" (ID:")[0]
                if customer_label != "Select Client"
                else "Select Client"
            )

        with col2:

            invoice_date = st.date_input(
                "Invoice Date",
                format="DD-MM-YYYY",
                key=f"sales_invoice_date_{v}"
            )

        col3, col4 = st.columns(2)

        with col3:

            if "current_invoice_number" not in st.session_state:
                st.session_state.current_invoice_number = generate_invoice_number()

            invoice_number = st.session_state.current_invoice_number

            st.text_input(
                "Invoice Number",
                value=invoice_number,
                disabled=True
            )

        with col4:

            due_date = st.date_input(
                "Due Date",
                format="DD-MM-YYYY",
                key=f"sales_due_date_{v}"
            )

        st.divider()

        # -------- MULTIPLE INVOICE ITEMS --------

        if "invoice_items" not in st.session_state:
            st.session_state.invoice_items = []

        st.subheader("Invoice Items")

        col1, col2, col3, col4 = st.columns(4)

        with col1:

            item_description = st.text_input(
                "Description",
                key=f"item_description_{v}"
            )

        with col2:

            quantity = st.number_input(
                "Quantity",
                min_value=1,
                value=1,
                key=f"invoice_quantity_{v}"
            )

        with col3:

            rate = st.number_input(
                "Rate",
                min_value=0.0,
                key=f"invoice_rate_{v}"
            )

        with col4:

            discount = st.number_input(
                "Discount",
                min_value=0.0,
                value=0.0,
                key=f"invoice_discount_{v}"
            )

        item_total = (quantity * rate) - discount

        st.info(
            f"Item Total: ${item_total:,.2f}"
        )

        if st.button(
            "➕ Add Item",
            key="add_invoice_item"
        ):

            if item_description.strip():

                st.session_state.invoice_items.append(
                    {
                        "Description": item_description,
                        "Quantity": quantity,
                        "Rate": rate,
                        "Discount": discount,
                        "Amount": item_total
                    }
                )

                for k in [
                    "item_description",
                    "invoice_quantity",
                    "invoice_rate",
                    "invoice_discount",
                ]:
                    if k in st.session_state:
                        del st.session_state[k]

                st.rerun()

            else:

                st.warning(
                    "Please enter description"
                )



        if st.session_state.invoice_items:

            item_df = pd.DataFrame(
                st.session_state.invoice_items
            )

            # recalculate amount
            item_df["Amount"] = (
                (
                    pd.to_numeric(item_df["Quantity"], errors="coerce").fillna(0)
                    *
                    pd.to_numeric(item_df["Rate"], errors="coerce").fillna(0)
                )
                -
                pd.to_numeric(item_df["Discount"], errors="coerce").fillna(0)
            )

            # add a Delete? column for selection
            item_df["Delete?"] = False

            edited_df = st.data_editor(
                item_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Description": st.column_config.TextColumn(
                        "Description"
                    ),
                    "Quantity": st.column_config.NumberColumn(
                        "Quantity",
                        min_value=1,
                        step=1
                    ),
                    "Rate": st.column_config.NumberColumn(
                        "Rate",
                        min_value=0.0,
                        step=0.01
                    ),
                    "Discount": st.column_config.NumberColumn(
                        "Discount",
                        min_value=0.0,
                        step=0.01
                    ),
                    "Amount": st.column_config.NumberColumn(
                        "Amount",
                        format="$%.2f",
                        disabled=True
                    ),
                    "Delete?": st.column_config.CheckboxColumn(
                        "Delete?",
                        default=False
                    )
                },
                key="invoice_items_editor"
            )

            # remove rows ticked for deletion
            if edited_df["Delete?"].any():
                if st.button("🗑️ Remove Selected", key="remove_selected_items"):
                    kept_df = edited_df[edited_df["Delete?"] == False].drop(columns=["Delete?"])
                    st.session_state.invoice_items = kept_df.to_dict("records")
                    st.rerun()

            # update session items with calculated amount (when nothing being deleted)
            st.session_state.invoice_items = (
                edited_df.drop(columns=["Delete?"]).to_dict("records")
            )

            amount = edited_df["Amount"].sum()

        else:
            amount = 0

        st.info(
            f"Service Amount: ${amount:,.2f}"
        )


        # -------- HST --------

        hst_rate = st.number_input(
            "HST Rate (%)",
            min_value=0.0,
            value=13.0,
            step=0.5,
            key=f"sales_hst_rate_{v}"
        )


        calculated_hst = amount * (hst_rate / 100)


        tax = st.number_input(
            "HST Amount",
            min_value=0.0,
            value=float(calculated_hst),
            step=0.01,
            key=f"sales_tax_{v}"
        )


        total = amount + tax

        # -------- TOTAL AMOUNT DUE --------

        st.info(
            f"Service Amount: ${amount:,.2f}"
        )


        st.warning(
            f"HST: ${tax:,.2f}"
        )


        st.success(
            f"Total Amount Due: ${total:,.2f}"
        )


        st.divider()



        # -------- PAYMENT STATUS --------

        payment_status = st.selectbox(
            "Status",
            [
                "Unpaid",
                "Paid"
            ],
            key=f"sales_payment_status_{v}"
        )


        if payment_status == "Paid":

            received_date = st.date_input(
                "Payment Received Date"
            )

        else:

            received_date = None



        st.divider()

    # -------- GENERATE INVOICE --------
        if st.button("🧾 Generate Invoice"):
            st.write("BUTTON CLICKED")   # TEMPORARY TEST LINE
            # ================= TOTAL =================
            amount = sum(
                item["Amount"]
                for item in st.session_state.invoice_items
            )
            total = amount + tax

            # -------- SAVE DATABASE --------
            total_quantity = sum(
                item["Quantity"]
                for item in st.session_state.invoice_items
            )
            item_descriptions = ", ".join(
                item["Description"]
                for item in st.session_state.invoice_items
            )
            add_invoice(
                invoice_number,
                customer_name,
                invoice_date,
                due_date,
                item_descriptions,
                total_quantity,
                0,
                amount,
                tax,
                total,
                payment_status,
                received_date
            )
            add_invoice_items(
                invoice_number,
                st.session_state.invoice_items
            )
            st.success(
                "Invoice saved successfully ✅"
            )

            # ---- reset everything for next invoice ----
            st.session_state.invoice_items = []
            st.session_state.current_invoice_number = generate_invoice_number()
            st.session_state.sales_form_version += 1

            st.rerun()

    # -------- UNPAID INVOICES --------

    st.divider()

    st.subheader("📌 Unpaid Invoices")


    invoices = get_invoices()


    unpaid_invoices = [
        inv
        for inv in invoices
        if inv["payment_status"] == "Unpaid"
        and not inv.get("is_voided")
    ]


    if unpaid_invoices:
        unpaid_df = pd.DataFrame(
            unpaid_invoices
        )

        total_unpaid = unpaid_df["total"].sum()
        st.metric("Total Unpaid", f"${total_unpaid:,.2f}")

        st.dataframe(
            unpaid_df[
                [
                    "invoice_number",
                    "client_name",
                    "invoice_date",
                    "due_date",
                    "total",
                    "payment_status",
                    "created_by"
                ]
            ].rename(columns={"created_by": "Created By"}),
            use_container_width=True,
            hide_index=True
        )


        st.markdown("**Mark Invoice as Paid**")

        unpaid_numbers = [
            inv["invoice_number"]
            for inv in unpaid_invoices
        ]

        mark_paid_invoice = st.selectbox(
            "Select Invoice",
            ["Select Invoice"] + unpaid_numbers,
            key="mark_paid_select"
        )

        if mark_paid_invoice != "Select Invoice" and can_edit():
            payment_received_date = st.date_input(
                "Payment Received Date",
                key="mark_paid_date"
            )
            if st.button("✅ Mark as Paid", key="mark_paid_button"):
                mark_invoice_paid(mark_paid_invoice, payment_received_date)
                st.success("Invoice marked as Paid")
                st.rerun()


    else:

        st.info(
            "No unpaid invoices"
        )
        # -------- AGING ANALYSIS --------
    st.divider()
    st.subheader("📅 Aging Analysis")
    if unpaid_invoices:
        aging_df = pd.DataFrame(unpaid_invoices)
        aging_df["due_date"] = pd.to_datetime(aging_df["due_date"], errors="coerce")
        today = pd.Timestamp.now().normalize()
        aging_df["days_overdue"] = (today - aging_df["due_date"]).dt.days
        aging_df = aging_df[aging_df["days_overdue"] > 0]
        def bucket_label(days):
            if days <= 10:
                return "0-10 Days"
            elif days <= 30:
                return "10-30 Days"
            elif days <= 90:
                return "30-90 Days"
            else:
                return "90+ Days"
        aging_df["Bucket"] = aging_df["days_overdue"].apply(bucket_label)
        bucket_order = ["0-10 Days", "10-30 Days", "30-90 Days", "90+ Days"]
        bucket_summary = (
            aging_df.groupby("Bucket")["total"]
            .sum()
            .reindex(bucket_order, fill_value=0)
        )
        col_a, col_b, col_c, col_d = st.columns(4)
        with col_a:
            st.metric("0-10 Days", f"${bucket_summary['0-10 Days']:,.2f}")
        with col_b:
            st.metric("10-30 Days", f"${bucket_summary['10-30 Days']:,.2f}")
        with col_c:
            st.metric("30-90 Days", f"${bucket_summary['30-90 Days']:,.2f}")
        with col_d:
            st.metric("90+ Days", f"${bucket_summary['90+ Days']:,.2f}")

        aging_display_df = aging_df[
            [
                "invoice_number",
                "client_name",
                "invoice_date",
                "due_date",
                "days_overdue",
                "Bucket",
                "total"
            ]
        ].sort_values("days_overdue", ascending=False)

        aging_display_df["invoice_date"] = pd.to_datetime(
            aging_display_df["invoice_date"], errors="coerce"
        ).dt.strftime("%Y-%m-%d")
        aging_display_df["due_date"] = aging_display_df["due_date"].dt.strftime("%Y-%m-%d")

        st.dataframe(
            aging_display_df,
            use_container_width=True,
            hide_index=True
        )

        aging_grand_total = aging_display_df["total"].sum()
        st.markdown(f"**Total Overdue Amount: ${aging_grand_total:,.2f}**")

        # -------- DOWNLOAD AGING SUMMARY --------

        aging_excel_df = aging_display_df.rename(columns={
            "invoice_number": "Invoice Number",
            "client_name": "Client Name",
            "invoice_date": "Invoice Date",
            "due_date": "Due Date",
            "days_overdue": "Days Overdue",
            "Bucket": "Aging Bucket",
            "total": "Total"
        })

        total_row = pd.DataFrame([{
            "Invoice Number": "",
            "Client Name": "",
            "Invoice Date": "",
            "Due Date": "",
            "Days Overdue": "",
            "Aging Bucket": "TOTAL",
            "Total": aging_grand_total
        }])

        aging_excel_df = pd.concat([aging_excel_df, total_row], ignore_index=True)

        excel_buffer = io.BytesIO()

        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            aging_excel_df.to_excel(writer, index=False, sheet_name="Aging Summary")

        st.download_button(
            label="📥 Download Aging Summary (Excel)",
            data=excel_buffer.getvalue(),
            file_name=f"invoice_aging_summary_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_aging_excel"
        )

        if aging_df.empty:
            st.info("No overdue invoices 🎉")

    else:

        st.info("No unpaid invoices to analyze")



    # ================= INVOICE HISTORY =================


if page == "📄 Invoice History":


    st.title(
        "📄 Invoice History"
    )


    invoices = get_invoices()


    if invoices:


        invoice_df = pd.DataFrame(
            invoices
        )

        if "id" in invoice_df.columns:
            invoice_df = invoice_df.drop(columns=["id"])

        invoice_df.insert(0, "Sr. No", range(1, len(invoice_df) + 1))

        invoice_df = invoice_df.drop(columns=["quantity", "rate","client_id"])

        invoice_df = invoice_df.rename(columns={
            "invoice_number": "Invoice Number",
            "client_name": "Client Name",
            "invoice_date": "Invoice Date",
            "due_date": "Due Date",
            "description": "Description",
            "amount": "Amount",
            "tax": "Tax",
            "total": "Total",
            "payment_status": "Payment Status",
            "received_date": "Received Date",
            "created_at": "Created At",
            "is_voided": "Voided"
        })

        if "Voided" in invoice_df.columns:
            invoice_df["Voided"] = invoice_df["Voided"].apply(
                lambda x: "🚫 VOID" if x else ""
            )

        display_invoice_df = invoice_df.copy()

        for col in ["Amount", "Tax", "Total"]:
            display_invoice_df[col] = display_invoice_df[col].apply(format_amount)

        def highlight_voided(row):
            if row.get("Voided") == "🚫 VOID":
                return ["background-color: #fdeaea"] * len(row)
            return [""] * len(row)

        styled_invoice_df = display_invoice_df.style.apply(highlight_voided, axis=1)

        st.dataframe(
            styled_invoice_df,
            use_container_width=True,
            hide_index=True
        )

        invoice_excel = io.BytesIO()

        with pd.ExcelWriter(invoice_excel, engine="openpyxl") as writer:
            invoice_df.to_excel(writer, index=False, sheet_name="Invoices")

        invoice_excel.seek(0)

        st.download_button(
            "⬇️ Export Invoice History",
            data=invoice_excel,
            file_name="Invoice_History.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


        if is_admin():

            pending_void_requests = [
                inv for inv in invoices
                if inv.get("void_requested") and not inv.get("is_voided")
            ]

            if pending_void_requests:

                st.divider()

                st.subheader(f"⏳ Pending Void Requests ({len(pending_void_requests)})")

                for req in pending_void_requests:

                    with st.container(border=True):
                        st.write(f"**Invoice:** {req['invoice_number']} | **Client:** {req['client_name']}")
                        st.write(f"**Requested by:** {req.get('void_requested_by', '')}")
                        st.write(f"**Reason:** {req.get('void_requested_reason', '')}")

                        c1, c2 = st.columns(2)

                        with c1:
                            if st.button("✅ Approve & Void", key=f"quick_approve_{req['invoice_number']}"):
                                void_invoice(
                                    req['invoice_number'],
                                    req.get('void_requested_reason', '')
                                )
                                st.success(f"Invoice {req['invoice_number']} voided")
                                st.rerun()

                        with c2:
                            if st.button("❌ Reject", key=f"quick_reject_{req['invoice_number']}"):
                                reject_void_request(req['invoice_number'])
                                st.info("Request rejected")
                                st.rerun()

        st.divider()


        st.subheader(
            "Delete & Download Invoice"
        )


        invoice_numbers = [
            row["invoice_number"]
            for row in invoices
        ]


        selected_invoice = st.selectbox(
            "Select Invoice",
            ["Select Invoice"] + invoice_numbers
        )

        if selected_invoice != "Select Invoice":

            pdf_buffer = generate_invoice_pdf(selected_invoice)

            st.download_button(
                label="⬇️ Download Invoice PDF",
                data=pdf_buffer,
                file_name=f"Invoice_{selected_invoice}.pdf",
                mime="application/pdf"
            )

        if selected_invoice != "Select Invoice" and can_edit():

            st.divider()

            st.markdown("**🚫 Void Invoice**")

            current_invoice_info = next(
                inv for inv in invoices
                if inv["invoice_number"] == selected_invoice
            )

            already_voided = current_invoice_info.get("is_voided")
            already_requested = current_invoice_info.get("void_requested")

            if already_voided:

                st.warning(
                    f"This invoice is already VOID.\n\n"
                    f"Requested by: {current_invoice_info.get('void_requested_by', 'N/A')}\n\n"
                    f"Reason: {current_invoice_info.get('void_reason', '')}\n\n"
                    f"Voided by: {current_invoice_info.get('voided_by', '')} on {current_invoice_info.get('voided_at', '')}"
                )

            elif already_requested and is_admin():

                st.info(
                    f"⏳ Void request pending your approval\n\n"
                    f"Requested by: {current_invoice_info.get('void_requested_by', '')}\n\n"
                    f"Reason: {current_invoice_info.get('void_requested_reason', '')}"
                )

                c1, c2 = st.columns(2)

                with c1:
                    if st.button("✅ Approve & Void", key="approve_void_button"):
                        void_invoice(
                            selected_invoice,
                            current_invoice_info.get("void_requested_reason", "")
                        )
                        st.success(f"Invoice {selected_invoice} has been voided")
                        st.rerun()

                with c2:
                    if st.button("❌ Reject Request", key="reject_void_button"):
                        reject_void_request(selected_invoice)
                        st.info("Void request rejected")
                        st.rerun()

            elif already_requested and not is_admin():

                st.info(
                    f"⏳ Void request submitted, waiting for admin approval.\n\n"
                    f"Requested by: {current_invoice_info.get('void_requested_by', '')}\n\n"
                    f"Reason: {current_invoice_info.get('void_requested_reason', '')}"
                )

            elif is_admin():

                void_reason = st.text_area(
                    "Reason for voiding this invoice (required)",
                    key="void_reason_input"
                )

                if st.button("🚫 Void This Invoice", key="void_invoice_button"):

                    if void_reason.strip():
                        void_invoice(selected_invoice, void_reason)
                        st.success(f"Invoice {selected_invoice} has been voided")
                        st.rerun()
                    else:
                        st.warning("Please enter a reason before voiding")

            else:

                void_request_reason = st.text_area(
                    "Reason for requesting void (required)",
                    key="void_request_reason_input"
                )

                if st.button("📨 Request Void", key="request_void_button"):

                    if void_request_reason.strip():
                        request_void_invoice(selected_invoice, void_request_reason)
                        st.success("Void request sent to admin for approval")
                        st.rerun()
                    else:
                        st.warning("Please enter a reason before requesting")

        if is_admin():

            if "confirm_invoice_delete" not in st.session_state:
                st.session_state.confirm_invoice_delete = False
            if st.button(
                "🗑️ Delete Invoice"
            ):
                if selected_invoice != "Select Invoice":
                    st.session_state.confirm_invoice_delete = True
                else:
                    st.warning(
                        "Please select an invoice"
                    )
            if st.session_state.confirm_invoice_delete:
                st.warning(
                    f"⚠️ Are you sure you want to delete invoice {selected_invoice}?"
                )
                c1, c2 = st.columns(2)
                with c1:
                    if st.button(
                        "✅ Yes, Delete"
                    ):
                        delete_invoice(
                            selected_invoice
                        )
                        st.session_state.confirm_invoice_delete = False
                        st.success(
                            "Invoice deleted successfully"
                        )
                        st.rerun()
                with c2:
                    if st.button(
                        "❌ Cancel"
                    ):
                        st.session_state.confirm_invoice_delete = False
                        st.rerun()



    else:


        st.info(
            "No invoices created yet"
        )



# ================= MAIN =================

if page == "📊 Reports":

    # ---------------- CLIENT SELECTOR ----------------

    reports_clients = get_clients()

    reports_client_names = [c["client_name"] for c in reports_clients]

    selected_report_client = st.selectbox(
        "Select Client",
        ["Select Client"] + reports_client_names,
        key="reports_client_select"
    )

    if selected_report_client == "Select Client":

        st.info("👆 Please select a client before uploading statements.")

        st.stop()

    st.markdown(
        f"<p style='font-size:16px; color:#1f4e79;'>📁 Working on reports for: <b>{selected_report_client}</b></p>",
        unsafe_allow_html=True
    )

    st.divider()

    # ---------------- UPLOADER RESET COUNTERS ----------------

    if "scotia_uploader_version" not in st.session_state:
        st.session_state.scotia_uploader_version = 0

    if "visa_uploader_version" not in st.session_state:
        st.session_state.visa_uploader_version = 0

    uploaded_excel = st.file_uploader(
        "Upload Excel File",
        type=["xlsx"]
    )

    # ---------------- SCOTIA UPLOADER ----------------

    scotia_col1, scotia_col2 = st.columns([5, 1])

    with scotia_col1:
        with st.container(key="uploader_scotia_container"):
            uploaded_pdf = st.file_uploader(
                "Upload Scotia Bank Statement PDF(s)",
                type=["pdf"],
                accept_multiple_files=True,
                key=f"scotia_pdf_uploader_{st.session_state.scotia_uploader_version}"
            )

    with scotia_col2:
        st.write("")
        st.write("")
        if st.button("🗑️ Clear All", key="clear_scotia_files"):
            st.session_state.scotia_uploader_version += 1
            st.rerun()

    # ---------------- VISA UPLOADER ----------------

    visa_col1, visa_col2 = st.columns([5, 1])

    with visa_col1:
        with st.container(key="uploader_visa_container"):
            uploaded_visa_pdf = st.file_uploader(
                "Upload Visa - 6023 or 7866 Statement PDF(s)",
                type=["pdf"],
                accept_multiple_files=True,
                key=f"visa_pdf_uploader_{st.session_state.visa_uploader_version}"
            )

    with visa_col2:
        st.write("")
        st.write("")
        if st.button("🗑️ Clear All", key="clear_visa_files"):
            st.session_state.visa_uploader_version += 1
            st.rerun()

    # ---------------- TRIANGLE MASTERCARD UPLOADER ----------------

    if "triangle_uploader_version" not in st.session_state:
        st.session_state.triangle_uploader_version = 0

    triangle_col1, triangle_col2 = st.columns([5, 1])

    with triangle_col1:
        with st.container(key="uploader_triangle_container"):
            uploaded_triangle_pdf = st.file_uploader(
                "Upload Triangle Mastercard 7133 Statements PDF(s)",
                type=["pdf"],
                accept_multiple_files=True,
                key=f"triangle_pdf_uploader_{st.session_state.triangle_uploader_version}"
            )

    with triangle_col2:
        st.write("")
        st.write("")
        if st.button("🗑️ Clear All", key="clear_triangle_files"):
            st.session_state.triangle_uploader_version += 1
            st.rerun()

    df = None
    scotia_df = None
    visa_df = None
    triangle_df = None

    client_file_prefix = re.sub(r"[^A-Za-z0-9_\-]+", "_", selected_report_client).strip("_")

    # ---------------- EXCEL (mutually exclusive with PDF uploaders) ----------------

    if uploaded_excel is not None:

        df = pd.read_excel(uploaded_excel)

        # remove blank Excel columns
        df = df.loc[:, ~df.columns.astype(str).str.contains("^Unnamed")]

    else:

        # ---------------- SCOTIA PDF ----------------

        if uploaded_pdf:

            import pdfplumber

            st.success(f"{len(uploaded_pdf)} Scotia statement PDF(s) uploaded successfully")

            transactions = []

            for pdf_file in uploaded_pdf:

                with pdfplumber.open(pdf_file) as pdf:

                    current = None

                    for page in pdf.pages:
                        started = False

                        words = page.extract_words()

                        rows = {}

                        for w in words:
                            y = round(float(w["top"]), 1)

                            if y not in rows:
                                rows[y] = []

                            rows[y].append(w)

                        for y in sorted(rows):

                            line_words = sorted(
                                rows[y],
                                key=lambda x: float(x["x0"])
                            )

                            text = " ".join(
                                w["text"] for w in line_words
                            )

                            header_text = text.upper()

                            # start transaction table
                            if (
                                "DATE" in header_text
                                and "DESCRIPTION" in header_text
                                and (
                                    "DEBIT" in header_text
                                    or "WITHDRAW" in header_text
                                )
                            ):
                                started = True
                                continue

                            if not started:
                                continue

                            # ignore bottom summary
                            if (
                                "NO. OF DEBITS" in header_text
                                or "NO. OF CREDITS" in header_text
                                or "TOTAL AMOUNT" in header_text
                                or "PAGE -" in header_text
                            ):
                                continue

                            # ignore repeated headers
                            if (
                                "DATE" in header_text
                                and "DESCRIPTION" in header_text
                            ):
                                continue

                            first = line_words[0]["text"]

                            # transaction row
                            if (
                                "/" in first
                                or "-" in first
                                or "." in first
                            ):

                                if current:
                                    transactions.append(current)

                                current = {
                                    "Date": first,
                                    "Description": "",
                                    "Debit": "",
                                    "Credit": "",
                                }

                                for w in line_words[1:]:

                                    x = float(w["x0"])
                                    value = w["text"]

                                    # amount columns based on PDF layout

                                    if x < 260:
                                        current["Description"] += " " + value

                                    elif x >= 260 and x < 400:
                                        current["Debit"] += " " + value

                                    elif x >= 400 and x < 540:
                                        current["Credit"] += " " + value

                            else:

                                # continuation description
                                if current:
                                    current["Description"] += " " + text

                    if current:
                        transactions.append(current)

            scotia_df = pd.DataFrame(transactions)

            if not scotia_df.empty:
                scotia_df = scotia_df.apply(
                    lambda x: x.str.strip()
                    if x.dtype == "object"
                    else x
                )

                scotia_df.columns = (
                    scotia_df.columns
                    .astype(str)
                    .str.replace("\n", " ", regex=False)
                    .str.strip()
                )

                for col in scotia_df.columns:

                    if "Deposit" in col or "Credit" in col:
                        scotia_df.rename(columns={col: "Credit"}, inplace=True)

                    if "Withdraw" in col or "Debit" in col:
                        scotia_df.rename(columns={col: "Debit"}, inplace=True)

                for col in ["Debit", "Credit"]:

                    if col in scotia_df.columns:

                        scotia_df[col] = (
                            scotia_df[col]
                            .astype(str)
                            .str.replace("$", "", regex=False)
                            .str.replace(",", "", regex=False)
                            .str.strip()
                        )

                        scotia_df[col] = pd.to_numeric(
                            scotia_df[col],
                            errors="coerce"
                        ).fillna(0)

                # ---------------- SCOTIA RULES ----------------

                scotia_df.loc[
                    scotia_df["Credit"].notna() &
                    scotia_df["Description"].astype(str).str.contains(
                        "MISC PAYMENT|TRANSFER FROM|DEPOSIT|DEP. FROM ANOTHER PARTY|SCOTIABANK PAYMENT",
                        case=False
                    ),
                    "Category"
                ] = "Revenue"

                scotia_df.loc[
                    scotia_df["Credit"].notna() &
                    scotia_df["Description"].astype(str).str.contains(
                        "Insurance|HEALTH/DENTAL CLAIM",
                        case=False
                    ),
                    "Category"
                ] = "Other Income"

                scotia_df.loc[
                    scotia_df["Debit"].notna() &
                    scotia_df["Description"].astype(str).str.strip().str.lower().eq("misc payment"),
                    "Category"
                ] = "Misc Expenses"

                scotia_df.loc[
                    scotia_df["Debit"].notna() &
                    scotia_df["Description"].astype(str).str.contains("INSURANCE", case=False),
                    "Category"
                ] = "Insurance"

                scotia_df.loc[
                    scotia_df["Debit"].notna() &
                    scotia_df["Description"].astype(str).str.contains("LOANS", case=False),
                    "Category"
                ] = "Car Loan"

                scotia_df.loc[
                    scotia_df["Debit"].notna() &
                    scotia_df["Description"].astype(str).str.contains("PC Bill Payment", case=False),
                    "Category"
                ] = "Purchases"

                scotia_df.loc[
                    scotia_df["Debit"].notna() &
                    scotia_df["Description"].astype(str).str.contains("GOODLIFE FITNESS", case=False),
                    "Category"
                ] = "Personal Expenses"

                scotia_df.loc[
                    scotia_df["Debit"].notna() &
                    scotia_df["Description"].astype(str).str.contains("HIGHWAY", case=False),
                    "Category"
                ] = "Parking and Toll"

                scotia_df.loc[
                    scotia_df["Debit"].notna() &
                    scotia_df["Description"].astype(str).str.contains("TSCC", case=False),
                    "Category"
                ] = "Vehicle Expense"

                scotia_df.loc[
                    scotia_df["Debit"].notna() &
                    scotia_df["Description"].astype(str).str.contains("Debit Memo", case=False),
                    "Category"
                ] = "Ask from Customer"

                scotia_df.loc[
                    scotia_df["Debit"].notna() &
                    scotia_df["Description"].astype(str).str.contains("SERVICE CHARGE|FEE", case=False),
                    "Category"
                ] = "Interest and Bank charges"
            else:
                st.warning("No transactions found in the Scotia statement(s).")

        # ---------------- VISA PDF ----------------

        if uploaded_visa_pdf:

            st.success(f"{len(uploaded_visa_pdf)} Visa statement(s) uploaded successfully")

            visa_dfs = []

            for visa_file in uploaded_visa_pdf:
                parsed_visa = parse_visa_statement(visa_file)
                if not parsed_visa.empty:
                    visa_dfs.append(parsed_visa)

            if visa_dfs:
                visa_df = pd.concat(visa_dfs, ignore_index=True)
                visa_df = apply_visa_categories(visa_df)
            else:
                st.warning("No transactions found in the Visa statement(s).")

        # ---------------- TRIANGLE MASTERCARD PDF ----------------

        if uploaded_triangle_pdf:

            st.success(f"{len(uploaded_triangle_pdf)} Triangle Mastercard statement(s) uploaded successfully")

            triangle_dfs = []

            for triangle_file in uploaded_triangle_pdf:
                parsed_triangle = parse_triangle_statement(triangle_file)
                if not parsed_triangle.empty:
                    triangle_dfs.append(parsed_triangle)

            if triangle_dfs:
                triangle_df = pd.concat(triangle_dfs, ignore_index=True)
                triangle_df = apply_visa_categories(triangle_df)
            else:
                st.warning("No transactions found in the Triangle Mastercard statement(s).")

    # ---------------- CLEAN DATA (Excel path only; PDF sources are cleaned above) ----------------

    any_upload = (
        uploaded_excel is not None
        or uploaded_pdf
        or uploaded_visa_pdf
        or uploaded_triangle_pdf
    )

    if not any_upload:

        st.markdown("""
        <div style="
            background:#f5f9ff;
            padding:25px;
            border-radius:20px;
            text-align:center;
            border:1px solid #d6e4f0;
        ">

        <h2 style="color:#1f4e79;">
        👋 Welcome to Prime Accounting
        </h2>

        <p style="font-size:18px;color:#555;">
        Upload a bank statement or credit card statement
        to start automated categorization and reporting
        </p>

        </div>
        """, unsafe_allow_html=True)

    if uploaded_excel is not None:

        if df is None or df.empty:
            st.warning("No transactions found in the uploaded file(s).")
            st.stop()

        df.columns = df.columns.astype(str).str.strip()

        for col in ["Debit", "Credit"]:

            if col in df.columns:

                df[col] = (
                    df[col]
                    .astype(str)
                    .str.replace("$", "", regex=False)
                    .str.replace(",", "", regex=False)
                    .str.strip()
                )

                df[col] = pd.to_numeric(
                    df[col],
                    errors="coerce"
                ).fillna(0)

        # ---------------- RULES ----------------

        df.loc[
            df["Credit"].notna() &
            df["Description"].astype(str).str.contains(
                "MISC PAYMENT|TRANSFER FROM|DEPOSIT|DEP. FROM ANOTHER PARTY|SCOTIABANK PAYMENT",
                case=False
            ),
            "Category"
        ] = "Revenue"

        df.loc[
            df["Credit"].notna() &
            df["Description"].astype(str).str.contains(
                "Insurance|HEALTH/DENTAL CLAIM",
                case=False
            ),
            "Category"
        ] = "Other Income"

        df.loc[
            df["Debit"].notna() &
            df["Description"].astype(str).str.strip().str.lower().eq("misc payment"),
            "Category"
        ] = "Misc Expenses"

        df.loc[
            df["Debit"].notna() &
            df["Description"].astype(str).str.contains("INSURANCE", case=False),
            "Category"
        ] = "Insurance"

        df.loc[
            df["Debit"].notna() &
            df["Description"].astype(str).str.contains("LOANS", case=False),
            "Category"
        ] = "Car Loan"

        df.loc[
            df["Debit"].notna() &
            df["Description"].astype(str).str.contains("PC Bill Payment", case=False),
            "Category"
        ] = "Purchases"

        df.loc[
            df["Debit"].notna() &
            df["Description"].astype(str).str.contains("GOODLIFE FITNESS", case=False),
            "Category"
        ] = "Personal Expenses"

        df.loc[
            df["Debit"].notna() &
            df["Description"].astype(str).str.contains("HIGHWAY", case=False),
            "Category"
        ] = "Parking and Toll"

        df.loc[
            df["Debit"].notna() &
            df["Description"].astype(str).str.contains("TSCC", case=False),
            "Category"
        ] = "Vehicle Expense"

        df.loc[
            df["Debit"].notna() &
            df["Description"].astype(str).str.contains("Debit Memo", case=False),
            "Category"
        ] = "Ask from Customer"

        df.loc[
            df["Debit"].notna() &
            df["Description"].astype(str).str.contains("SERVICE CHARGE|FEE", case=False),
            "Category"
        ] = "Interest and Bank charges"

    # =====================================================================
    # REUSABLE RENDER: transaction table + category summary for one source
    # (P&L is intentionally NOT generated here — see combined P&L below)
    # =====================================================================

    def render_source_section(source_df, source_label, file_key):

        source_df = source_df.reset_index(drop=True)

        source_df.insert(0, "Sr. No", range(1, len(source_df) + 1))

        # ---------------- DISPLAY TABLE ----------------
        display_source_df = source_df.copy()

        for col in ["Credit", "Debit"]:
            if col in display_source_df.columns:
                display_source_df[col] = display_source_df[col].apply(format_amount)

        st.subheader(f"📊 {source_label} — Categorized Transactions")

        # ---------------- TOTAL ROW ----------------

        debit_total = pd.to_numeric(source_df["Debit"], errors="coerce").fillna(0).sum()
        credit_total = pd.to_numeric(source_df["Credit"], errors="coerce").fillna(0).sum()

        total_row = pd.DataFrame([{
            "Sr. No": "",
            "Date": "",
            "Description": "TOTAL",
            "Debit": debit_total,
            "Credit": credit_total
        }])

        display_source_df = pd.concat([display_source_df, total_row], ignore_index=True)

        for col in ["Credit", "Debit"]:
            if col in display_source_df.columns:
                display_source_df[col] = display_source_df[col].apply(format_amount)

        def style_transaction_row(row):

            if row["Description"] == "TOTAL":
                return ["font-weight: bold"] * len(row)

            if row.get("Needs Review", False):
                return ["background-color: #fff3cd"] * len(row)

            return [""] * len(row)

        styled_transactions = display_source_df.style.apply(style_transaction_row, axis=1)

        st.dataframe(styled_transactions, use_container_width=True, hide_index=True)

        # ---------------- CATEGORY STATUS COUNT ----------------

        total_entries = len(source_df)

        categorized_entries = (
            source_df["Category"]
            .fillna("")
            .astype(str)
            .str.strip()
            .ne("")
            .sum()
        )

        uncategorized_entries = total_entries - categorized_entries

        st.markdown(f"**📌 {source_label} Categorization Summary**")
        st.markdown(f"""
        - ✅ **Total Transactions:** {total_entries:,}
        - 🟢 **Categorized Transactions:** {categorized_entries:,}
        - ⚪ **Uncategorized Transactions:** {uncategorized_entries:,}
        """)

        # ---------------- DOWNLOAD TRANSACTIONS (with highlight + client header) ----------------
        from openpyxl.styles import PatternFill

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            source_df.to_excel(writer, index=False, sheet_name="Transactions", startrow=1)

            worksheet = writer.sheets["Transactions"]
            worksheet.cell(row=1, column=1).value = f"Client: {selected_report_client}  |  Source: {source_label}"

            highlight_fill = PatternFill(
                start_color="FFF3CD",
                end_color="FFF3CD",
                fill_type="solid"
            )

            if "Needs Review" in source_df.columns:
                for row_num, needs_review in enumerate(source_df["Needs Review"], start=3):
                    if needs_review:
                        for col_num in range(1, len(source_df.columns) + 1):
                            worksheet.cell(row=row_num, column=col_num).fill = highlight_fill

        output.seek(0)

        st.download_button(
            f"⬇️ Export {source_label} Categorized Data",
            data=output,
            file_name=f"{client_file_prefix}_{source_label.replace(' ', '_')}_Categorized_Data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_{file_key}_transactions"
        )

        # ---------------- CATEGORY SUMMARY ----------------

        source_df["Amount"] = source_df["Credit"].fillna(0) + source_df["Debit"].fillna(0)

        summary = source_df.groupby("Category")["Amount"].sum().reset_index()

        display_summary = summary.copy()
        display_summary["Amount"] = display_summary["Amount"].apply(format_amount)

        summary_total = summary["Amount"].sum()

        total_summary_row = pd.DataFrame([{
            "Category": "TOTAL",
            "Amount": format_amount(summary_total)
        }])

        display_summary = pd.concat([display_summary, total_summary_row], ignore_index=True)

        st.subheader(f"📋 {source_label} — Category Summary")

        def bold_total(row):
            if row["Category"] == "TOTAL":
                return ["font-weight: bold"] * len(row)
            return [""] * len(row)

        styled_summary = display_summary.style.apply(bold_total, axis=1)

        st.dataframe(styled_summary, use_container_width=True, hide_index=True)

        # ---------------- SUMMARY DOWNLOAD (with client header) ----------------

        summary_output = io.BytesIO()

        with pd.ExcelWriter(summary_output, engine="openpyxl") as writer:
            summary.to_excel(writer, index=False, sheet_name="Category Summary", startrow=1)
            worksheet = writer.sheets["Category Summary"]
            worksheet.cell(row=1, column=1).value = f"Client: {selected_report_client}  |  Source: {source_label}"

        summary_output.seek(0)

        st.download_button(
            f"⬇️ Export {source_label} Summary Data",
            data=summary_output,
            file_name=f"{client_file_prefix}_{source_label.replace(' ', '_')}_Category_Summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_{file_key}_summary"
        )

        st.divider()

        return source_df

    # =====================================================================
    # RENDER EACH UPLOADED SOURCE SEPARATELY
    # =====================================================================

    combined_pl_sources = []

    if uploaded_excel is not None and df is not None and not df.empty:
        df = render_source_section(df, "Excel Upload", "excel")
        combined_pl_sources.append(df)

    if scotia_df is not None and not scotia_df.empty:
        scotia_df = render_source_section(scotia_df, "Scotia Bank", "scotia")
        combined_pl_sources.append(scotia_df)

    if visa_df is not None and not visa_df.empty:
        visa_df = render_source_section(visa_df, "Visa", "visa")
        combined_pl_sources.append(visa_df)

    if triangle_df is not None and not triangle_df.empty:
        triangle_df = render_source_section(triangle_df, "Triangle Mastercard", "triangle")
        combined_pl_sources.append(triangle_df)

    # =====================================================================
    # ONE COMBINED PROFIT & LOSS ACROSS ALL UPLOADED SOURCES
    # =====================================================================

    if combined_pl_sources:

        combined_df = pd.concat(
            [d[["Description", "Debit", "Credit", "Category"]] for d in combined_pl_sources],
            ignore_index=True
        )

        # drop categories that should never appear on the P&L (e.g. internal
        # credit card payments between the client's own accounts)
        combined_df = combined_df[~combined_df["Category"].isin(PL_EXCLUDED_CATEGORIES)]

        st.subheader(f"📊 Combined Profit & Loss Statement — {selected_report_client}")

        revenue_by_cat = (
            combined_df[combined_df["Category"].isin(PL_REVENUE_CATEGORIES)]
            .groupby("Category")["Credit"].sum()
        )

        cogs_by_cat = (
            combined_df[combined_df["Category"].isin(PL_COGS_CATEGORIES)]
            .groupby("Category")["Debit"].sum()
        )

        expense_by_cat = (
            combined_df[combined_df["Category"].isin(PL_EXPENSE_CATEGORIES)]
            .groupby("Category")["Debit"].sum()
        )

        known_categories = (
            set(PL_REVENUE_CATEGORIES)
            | set(PL_COGS_CATEGORIES)
            | set(PL_EXPENSE_CATEGORIES)
            | set(PL_EXCLUDED_CATEGORIES)
        )

        leftover_df = combined_df[~combined_df["Category"].isin(known_categories)]
        # also treat blank/uncategorized rows as "Other/Uncategorized"
        other_uncategorized = leftover_df["Debit"].sum() - leftover_df["Credit"].sum()

        total_revenue = sum(revenue_by_cat.get(c, 0) for c in PL_REVENUE_CATEGORIES)
        total_cogs = sum(cogs_by_cat.get(c, 0) for c in PL_COGS_CATEGORIES)
        gross_profit = total_revenue - total_cogs

        total_expenses = sum(expense_by_cat.get(c, 0) for c in PL_EXPENSE_CATEGORIES)
        if other_uncategorized:
            total_expenses += other_uncategorized

        net_income_before_taxes = gross_profit - total_expenses

        # ---------------- TAX RATE INPUTS ----------------

        tax_col1, tax_col2 = st.columns(2)

        with tax_col1:
            federal_tax_rate = st.number_input(
                "Federal Tax Rate (%)",
                min_value=0.0,
                value=0.0,
                step=0.5,
                key="pl_federal_tax_rate"
            )

        with tax_col2:
            provincial_tax_rate = st.number_input(
                "Provincial Tax Rate (%)",
                min_value=0.0,
                value=0.0,
                step=0.5,
                key="pl_provincial_tax_rate"
            )

        combined_tax_rate = federal_tax_rate + provincial_tax_rate

        tax_expense = (
            net_income_before_taxes * (combined_tax_rate / 100)
            if net_income_before_taxes > 0
            else 0
        )

        net_income_after_taxes = net_income_before_taxes - tax_expense

        # ---------------- BUILD P&L ROW LIST ----------------
        # Each row: (label, amount_or_None, row_type)
        # row_type drives styling: "section" (blue), "subtotal" (gray/bold),
        # "line" (plain), "rate" (plain, percentage)

        pl_rows = []

        pl_rows.append(("Revenue", None, "section"))
        for cat in PL_REVENUE_CATEGORIES:
            pl_rows.append((cat, revenue_by_cat.get(cat, 0), "line"))
        pl_rows.append(("Total Revenue", total_revenue, "subtotal"))

        pl_rows.append(("Cost of Goods Sold", None, "section"))
        for cat in PL_COGS_CATEGORIES:
            pl_rows.append((cat, cogs_by_cat.get(cat, 0), "line"))
        pl_rows.append(("Total Cost of Goods Sold", total_cogs, "subtotal"))

        pl_rows.append(("Gross Profit", gross_profit, "grossprofit"))

        pl_rows.append(("Expenses", None, "section"))
        for cat in PL_EXPENSE_CATEGORIES:
            pl_rows.append((cat, expense_by_cat.get(cat, 0), "line"))
        if other_uncategorized:
            pl_rows.append(("Other/Uncategorized", other_uncategorized, "line"))
        pl_rows.append(("Total Expenses", total_expenses, "subtotal"))

        pl_rows.append(("Net income before taxes", net_income_before_taxes, "grossprofit"))
        pl_rows.append(("Federal Tax rate", federal_tax_rate, "rate"))
        pl_rows.append(("Provincial Tax rate", provincial_tax_rate, "rate"))
        pl_rows.append(("Tax Expense", tax_expense, "grossprofit"))
        pl_rows.append(("Net income after taxes", net_income_after_taxes, "section_total"))

        pl_df = pd.DataFrame(pl_rows, columns=["Description", "Amount", "RowType"])

        # ---------------- DISPLAY ----------------

        pl_display = pl_df.copy()

        def format_pl_amount(row):
            if row["RowType"] == "rate":
                return f"{row['Amount']:.2f}%" if pd.notna(row["Amount"]) else ""
            return format_amount(row["Amount"]) if pd.notna(row["Amount"]) else ""

        pl_display["Amount"] = pl_display.apply(format_pl_amount, axis=1)
        pl_display = pl_display.drop(columns=["RowType"])

        def style_pl_row(row):
            row_type = pl_df.loc[row.name, "RowType"]
            if row_type == "section":
                return ["background-color: #1f4e79; color: white; font-weight: bold"] * len(row)
            if row_type == "section_total":
                return ["background-color: #1f4e79; color: white; font-weight: bold"] * len(row)
            if row_type == "grossprofit":
                return ["background-color: #e0e0e0; font-weight: bold"] * len(row)
            if row_type == "subtotal":
                return ["font-weight: bold; border-top: 1px solid #999"] * len(row)
            return [""] * len(row)

        styled_pl = pl_display.style.apply(style_pl_row, axis=1)

        st.dataframe(styled_pl, use_container_width=True, hide_index=True)

        # ---------------- COMBINED P&L DOWNLOAD (with client header + colored sections) ----------------

        from openpyxl.styles import PatternFill, Font

        pl_output = io.BytesIO()

        with pd.ExcelWriter(pl_output, engine="openpyxl") as writer:

            pl_df.drop(columns=["RowType"]).to_excel(
                writer, index=False, sheet_name="Profit & Loss", startrow=1
            )

            worksheet = writer.sheets["Profit & Loss"]
            worksheet.cell(row=1, column=1).value = (
                f"Client: {selected_report_client}  |  Combined Profit & Loss"
            )

            section_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            section_font = Font(color="FFFFFF", bold=True)

            grossprofit_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            bold_font = Font(bold=True)

            # data starts at Excel row 3 (row 1 = client header, row 2 = column headers)
            for i, row_type in enumerate(pl_df["RowType"]):
                excel_row = i + 3

                if row_type == "rate":
                    # write as a numeric percentage value in the Amount column
                    amount_value = pl_df.loc[i, "Amount"]
                    cell = worksheet.cell(row=excel_row, column=2)
                    cell.value = amount_value / 100 if pd.notna(amount_value) else None
                    cell.number_format = "0.00%"

                if row_type == "section":
                    for col_num in range(1, 3):
                        worksheet.cell(row=excel_row, column=col_num).fill = section_fill
                        worksheet.cell(row=excel_row, column=col_num).font = section_font

                elif row_type == "section_total":
                    for col_num in range(1, 3):
                        worksheet.cell(row=excel_row, column=col_num).fill = section_fill
                        worksheet.cell(row=excel_row, column=col_num).font = section_font

                elif row_type == "grossprofit":
                    for col_num in range(1, 3):
                        worksheet.cell(row=excel_row, column=col_num).fill = grossprofit_fill
                        worksheet.cell(row=excel_row, column=col_num).font = bold_font

                elif row_type == "subtotal":
                    for col_num in range(1, 3):
                        worksheet.cell(row=excel_row, column=col_num).font = bold_font


        pl_output.seek(0)

        st.download_button(
            "📊 Export Combined Profit & Loss Statement",
            data=pl_output,
            file_name=f"{client_file_prefix}_Profit_and_Loss.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_combined_pl"
        )



elif page == "🏠 Dashboard":

    col_logo, col_greeting = st.columns([1, 6])

    with col_logo:
        st.image("Logo.jpeg", width=80)

    with col_greeting:
        today_display = datetime.date.today().strftime("%A, %B %d, %Y")
        st.markdown(
            "<h1 style='color:#1f4e79; margin-bottom:0px;'>Prime Accounting and Tax</h1>",
            unsafe_allow_html=True
        )
        st.markdown(
            f"<p style='font-size:16px; color:#1f4e79; margin-top:0px;'>👋 Welcome back — Today is {today_display}</p>",
            unsafe_allow_html=True
        )

    st.markdown("""

    <style>

    .main-title {
        font-size:45px;
        font-weight:700;
        color:#1f4e79;
        margin-bottom:0px;
    }

    .sub-title {
        font-size:22px;
        color:#666;
        margin-top:0px;
    }

   .card {
        background:white;
        padding:25px;
        border-radius:18px;
        border:1px solid #e5e7eb;
        box-shadow:0 4px 15px rgba(0,0,0,0.08);
        text-align:center;
        height:140px;
        display:flex;
        flex-direction:column;
        justify-content:center;
    }
    .card-blue {
        background:#e8f1fb;
    }

    .card-amber {
        background:#fef3e0;
    }

    .card-red {
        background:#fdeaea;
    }

    .card-teal {
        background:#e3f4f4;
    }

    .card-title {
        font-size:18px;
        color:#666;
    }

    .card-number {
        font-size:35px;
        font-weight:bold;
        color:#1f4e79;
    }

    </style>
    """, unsafe_allow_html=True)


    st.write("")


    today = datetime.date.today()

    c1, c2, c3, c4 = st.columns(4)


    with c1:

        st.markdown(
            """
            <div class="card card-blue">
            <div class="card-title">
            👥 Clients
            </div>

            <div class="card-number">
            {}
            </div>

            </div>
            """.format(len(get_clients())),
            unsafe_allow_html=True
        )


    with c2:

        dashboard_invoices = get_invoices()

        dashboard_unpaid_count = len([
            inv for inv in dashboard_invoices
            if inv["payment_status"] == "Unpaid"
        ])

        st.markdown(
            """
            <div class="card card-amber">
            <div class="card-title">
            📌 Unpaid Invoices
            </div>
            <div class="card-number">
            {}
            </div>
            </div>
            """.format(dashboard_unpaid_count),
            unsafe_allow_html=True
        )


    with c3:

        overdue_30_invoices = [
            inv for inv in dashboard_invoices
            if inv["payment_status"] == "Unpaid"
            and inv.get("invoice_date")
            and (today - datetime.date.fromisoformat(inv["invoice_date"])).days > 30
        ]

        overdue_30_count = len(overdue_30_invoices)

        overdue_30_total = sum(
            inv["total"] for inv in overdue_30_invoices
        )

        st.markdown(
            """
            <div class="card card-red">

            <div class="card-title">
            ⚠️ 30+ Days Overdue
            </div>

            <div class="card-number">
            ${:,.2f}
            </div>

            </div>
            """.format(overdue_30_total),
            unsafe_allow_html=True
        )


    with c4:

        dashboard_unpaid_total = sum(
            inv["total"]
            for inv in dashboard_invoices
            if inv["payment_status"] == "Unpaid"
        )

        st.markdown(
            """
            <div class="card card-teal">

            <div class="card-title">
            💰 Unpaid Total
            </div>

            <div class="card-number">
            ${:,.2f}
            </div>

            </div>
            """.format(dashboard_unpaid_total),
            unsafe_allow_html=True
        )


    st.write("")

    st.divider()  


    st.subheader("🚀 Quick Actions")


    a,b,c = st.columns(3)


    with a:
        if st.button("➕ Add Client", use_container_width=True):
            st.session_state.page = "👥 Clients"
            st.session_state.clients_active_tab = "Add Client"
            st.rerun()


    with b:
        if st.button("📂 Upload Statement", use_container_width=True):
            st.session_state.page = "📊 Reports"
            st.rerun()


    with c:
        if st.button("🧾 Create Invoice", use_container_width=True):
            st.session_state.page = "🧾 Sales"
            st.rerun()


    st.divider()

    st.subheader("📊 Paid vs Unpaid Invoices")

    paid_total = sum(
        inv["total"] for inv in dashboard_invoices
        if inv["payment_status"] == "Paid"
    )

    unpaid_total_chart = sum(
        inv["total"] for inv in dashboard_invoices
        if inv["payment_status"] == "Unpaid"
    )

    if paid_total > 0 or unpaid_total_chart > 0:

        fig = go.Figure(data=[
            go.Pie(
                labels=["Paid", "Unpaid"],
                values=[paid_total, unpaid_total_chart],
                hole=0.5,
                marker=dict(colors=["#3a8fa3", "#d97b66"])
            )
        ])

        fig.update_layout(
            height=300,
            margin=dict(t=10, b=10, l=10, r=10)
        )

        st.plotly_chart(fig, use_container_width=True)

    else:

        st.info("No invoice data yet to display chart")


    st.divider()

    st.subheader("📅 Invoice Aging Summary")

    aging_dashboard_invoices = [
        inv for inv in dashboard_invoices
        if inv["payment_status"] == "Unpaid"
    ]

    if aging_dashboard_invoices:

        aging_dash_df = pd.DataFrame(aging_dashboard_invoices)

        aging_dash_df["due_date"] = pd.to_datetime(aging_dash_df["due_date"], errors="coerce")

        today_ts = pd.Timestamp.now().normalize()

        aging_dash_df["days_overdue"] = (today_ts - aging_dash_df["due_date"]).dt.days

        aging_dash_df = aging_dash_df[aging_dash_df["days_overdue"] > 0]

        if not aging_dash_df.empty:

            def aging_bucket(days):
                if days <= 10:
                    return "0-10 Days"
                elif days <= 30:
                    return "10-30 Days"
                elif days <= 90:
                    return "30-90 Days"
                else:
                    return "90+ Days"

            aging_dash_df["Bucket"] = aging_dash_df["days_overdue"].apply(aging_bucket)

            bucket_order = ["0-10 Days", "10-30 Days", "30-90 Days", "90+ Days"]

            aging_bucket_totals = (
                aging_dash_df.groupby("Bucket")["total"]
                .sum()
                .reindex(bucket_order, fill_value=0)
            )

            aging_fig = go.Figure(data=[
                go.Bar(
                    x=bucket_order,
                    y=aging_bucket_totals.values,
                    marker=dict(color=["#3a8fa3", "#f0ad4e", "#d9706b", "#a83232"]),
                    text=[f"${v:,.0f}" for v in aging_bucket_totals.values],
                    textposition="outside"
                )
            ])

            aging_fig.update_layout(
                height=320,
                margin=dict(t=20, b=10, l=10, r=10),
                xaxis_title="Days Overdue",
                yaxis_title="Amount ($)"
            )

            st.plotly_chart(aging_fig, use_container_width=True)

        else:

            st.info("No overdue invoices 🎉")

    else:

        st.info("No unpaid invoices to analyze")


    st.divider()


    st.subheader("⏰ Upcoming Due Dates (Next 7 Days)")


    today = datetime.date.today()
    week_later = today + datetime.timedelta(days=7)

    upcoming_invoices = [
        inv for inv in dashboard_invoices
        if inv["payment_status"] == "Unpaid"
        and inv.get("due_date")
        and today <= datetime.date.fromisoformat(inv["due_date"]) <= week_later
    ]

    if upcoming_invoices:

        upcoming_df = pd.DataFrame(upcoming_invoices)

        upcoming_df = upcoming_df[
            ["invoice_number", "client_name", "due_date", "total"]
        ]

        upcoming_df = upcoming_df.rename(columns={
            "invoice_number": "Invoice Number",
            "client_name": "Client Name",
            "due_date": "Due Date",
            "total": "Total"
        })

        upcoming_df["Total"] = upcoming_df["Total"].apply(format_amount)

        st.dataframe(
            upcoming_df,
            use_container_width=True,
            hide_index=True
        )

    else:

        st.info("No invoices due in the next 7 days")

st.markdown(
    """
    <div style="
        margin-top:40px;
        padding-top:12px;
        border-top:1px solid #e6e6e6;
        text-align:right;
        font-size:16px;
        color:#888;
    ">
        Prime Automated Categorization & Reporting System<br>
        <span style="font-weight:600;color:#1f4e79;">
            Powered by: Prime Accounting & Tax
        </span>
    </div>
    """,
    unsafe_allow_html=True
)
